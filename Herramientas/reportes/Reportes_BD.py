from PyQt5.QtWidgets import (QMainWindow, QMenu, QAction, QLineEdit, QPushButton, QLabel, QVBoxLayout, 
                             QHBoxLayout,QApplication, QWidget, QGroupBox,QSizePolicy, QSpacerItem,QFileDialog, QRadioButton,QProgressBar,QTextEdit, QComboBox, QMessageBox,QInputDialog,QDialog, QLabel, QLineEdit, QCheckBox, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox)
from PyQt5.QtCore import Qt,QThread, pyqtSignal
from PyQt5.QtGui import QIcon,QFont
import sys
from PyQt5.QtCore import QSettings
import os
import datetime
from PIL import Image 
import pandas as pd
import os
import psycopg2
import sqlite3
import uuid
import chardet
import subprocess
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill






from .codigos.consultas_reporte_1_bd import nombres_consultas_observaciones as nombres_consultas_obs
from .codigos.consultas_reporte_1_bd import consultas_reporte_1_bd as consultas_obs
from .codigos.consultas_reporte_2_mtj import consultas_reporte_2_mtj, nombres_consultas_observaciones_mtj
#from .codigos.crear_tablas import tablas_mtj
from .codigos.estandarizacion_mtj import ejecutar_estandarizacion
from .codigos.exportar_gdb import export_to_geopackage
from .codigos.exportar_gpk_2 import export_to_geopackage2
from .codigos.consultas_fili import ejecutar_consulta_fili
from .codigos.exportar_gpk_cons import export_to_geopackage3
from .codigos.estandarizacion_mtj_batch import open_estandarizacion_window

from .codigos.gpkgtoxtf import abrir_ventana_exportacion
from .codigos.xtfagpkg import abrir_ventana_importacion
from .codigos.xtfbd import abrir_ventana_importacionbd
from .codigos.bd_a_xtf import abrir_ventana_exportacion_bd_2



# Definir variable global root
root = None





class ReportGeneratorApp(QMainWindow):
    def __init__(self, iface):
        super(ReportGeneratorApp, self).__init__()
        self.iface = iface
        self.conexion = None
        self.setWindowTitle("Reportes de Validación BD-MTJ")
        self.setGeometry(100, 100, 600, 750)


        # Recuperar la última contraseña almacenada en QSettings
        settings = QSettings("ReportGeneratorApp", "BaseDatos")
        ultima_contraseña = settings.value("ultima_contraseña", "")

        # Crear widgets y menús
        self.create_widgets(ultima_contraseña)
        self.create_menu()

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Salir', "¿Está seguro que desea salir?", 
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()




    def create_widgets(self,ultima_contraseña):
        # Main layout
        main_layout = QVBoxLayout()

        # Group 1: Conexión a la base de datos
        db_groupbox = QGroupBox("Conexión a base de datos")
        db_layout = QVBoxLayout()

        # Host
        host_layout = QHBoxLayout()
        host_label = QLabel("Host de la base de datos:")
        self.entry_host = QLineEdit()
        self.entry_host.setText("localhost")
        host_layout.addWidget(host_label)
        host_layout.addWidget(self.entry_host)
        db_layout.addLayout(host_layout)

        # Port
        port_layout = QHBoxLayout()
        port_label = QLabel("Puerto de la base de datos:")
        self.entry_port = QLineEdit()
        self.entry_port.setText("5432")
        port_layout.addWidget(port_label)
        port_layout.addWidget(self.entry_port)
        db_layout.addLayout(port_layout)

        # Usuario
        user_layout = QHBoxLayout()
        user_label = QLabel("Usuario de la base de datos:")
        self.entry_usuario = QLineEdit()
        self.entry_usuario.setText("postgres")
        user_layout.addWidget(user_label)
        user_layout.addWidget(self.entry_usuario)
        db_layout.addLayout(user_layout)

        # Contraseña
        password_layout = QHBoxLayout()
        password_label = QLabel("Contraseña de la base de datos:")
        self.entry_password = QLineEdit()
        self.entry_password.setEchoMode(QLineEdit.Password)

        # Asignar la última contraseña recuperada de QSettings
        self.entry_password.setText(ultima_contraseña)

        password_layout.addWidget(password_label)
        password_layout.addWidget(self.entry_password)
        db_layout.addLayout(password_layout)

        # Conectar button
        connect_button_layout = QHBoxLayout()
        self.boton_conectar = QPushButton("Conectar")
        self.boton_conectar.clicked.connect(self.conectar_y_obtener_esquemas)
        connect_button_layout.addWidget(self.boton_conectar)
        db_layout.addLayout(connect_button_layout)

        # Nombre de la base de datos
        db_name_layout = QHBoxLayout()
        db_label = QLabel("Nombre de la base de datos:")
        self.db_menu = QComboBox()
        db_name_layout.addWidget(db_label)
        db_name_layout.addWidget(self.db_menu)

        # "+" button for creating a new database
        #create_db_button = QPushButton("+")
        #create_db_button.clicked.connect(self.crear_base_datos)
        #db_name_layout.addWidget(create_db_button)
        db_layout.addLayout(db_name_layout)

        # Esquema
        schema_layout = QHBoxLayout()
        schema_label = QLabel("Seleccione el esquema:")
        self.esquema_menu = QComboBox()
        schema_layout.addWidget(schema_label)
        schema_layout.addWidget(self.esquema_menu)

        # "+" button for creating a new schema
        #create_schema_button = QPushButton("+")
        #create_schema_button.clicked.connect(self.crear_esquema)
        #schema_layout.addWidget(create_schema_button)
        db_layout.addLayout(schema_layout)

        db_groupbox.setLayout(db_layout)
        main_layout.addWidget(db_groupbox)









        # Group 2: Seleccionar tipo de reporte
        report_frame = QGroupBox("Seleccionar Tipo de reporte")
        report_layout = QVBoxLayout()  # Mantener el QVBoxLayout para la disposición vertical

        # Crear un layout horizontal para el selector de reporte y el botón de generar
        top_layout = QHBoxLayout()

        # Selector de reporte
        report_label = QLabel("Seleccione el tipo de reporte:")
        self.reporte_menu = QComboBox()
        self.reporte_menu.addItems(["Observaciones BD", "BD MTJ Observaciones", "Unificado"])

        # Botón para generar el reporte
        self.generate_report_button = QPushButton("Generar Reporte")
        self.generate_report_button.clicked.connect(self.ejecutar_consultas)

        # Añadir el selector de reporte y el botón al layout horizontal
        top_layout.addWidget(report_label)
        top_layout.addWidget(self.reporte_menu)
        top_layout.addWidget(self.generate_report_button)

        # Añadir el layout superior al layout principal de reporte
        report_layout.addLayout(top_layout)

        # Crear un layout horizontal para "Formato" y los radio buttons
        formato_layout = QHBoxLayout()

        # Etiqueta "Formato"
        formato_label = QLabel("Formato:")

        # Añadir los radio buttons para Consolidado y Simplificado
        self.consolidado_radio = QRadioButton("Consolidado")
        self.simplificado_radio = QRadioButton("Simplificado")

        # Por defecto seleccionamos Consolidado
        self.consolidado_radio.setChecked(True)

        # Configurar el layout horizontal para ser compacto
        formato_layout.setSpacing(5)  # Reducir al mínimo el espacio entre los radio buttons
        formato_layout.setContentsMargins(0, 0, 0, 0)  # Sin márgenes entre los elementos

        # Añadir la etiqueta "Formato" y los radio buttons al layout horizontal
        formato_layout.addWidget(formato_label)
        formato_layout.addWidget(self.consolidado_radio)
        formato_layout.addWidget(self.simplificado_radio)

        # Añadir el layout "Formato" al layout principal de reporte (debajo del selector)
        report_layout.addLayout(formato_layout)

        # Ajustar los márgenes para que el contenido no esté tan separado
        formato_layout.setContentsMargins(0, 0, 0, 0)  # Sin márgenes

        report_frame.setLayout(report_layout)

        # Añadir el report_frame al layout principal
        main_layout.addWidget(report_frame)








        # Cuadro de texto para mostrar el estado detallado del progreso
        self.status_text = QTextEdit()
        self.status_text.setReadOnly(True)
        main_layout.addWidget(self.status_text)

        # Progress section
        #self.progress_bar = QProgressBar()
        #self.progress_bar.setValue(0)  # Iniciar en 0%
        #self.progress_bar.setMinimum(0)  # Valor mínimo
        #self.progress_bar.setMaximum(100)  # Valor máximo (100%)
        #main_layout.addWidget(self.progress_bar)

        # Etiquetas para autor e información
        autor_font = QFont()
        autor_font.setPointSize(7)
        info_frame = QHBoxLayout()

        sig_label = QLabel("SIG - SPO")
        sig_label.setFont(autor_font)
        info_frame.addWidget(sig_label, alignment=Qt.AlignRight)

        #author_label = QLabel("@jmariorod")
        #author_label.setFont(autor_font)
        #info_frame.addWidget(author_label, alignment=Qt.AlignRight)

        version_label = QLabel("Reportes V1.4.15.09.2024")
        version_label.setFont(autor_font)
        info_frame.addWidget(version_label, alignment=Qt.AlignRight)

        main_layout.addLayout(info_frame)

        # Main widget to hold the layout
        main_widget = QWidget()
        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)




    def update_progress(self, status):
        self.status_text.append(status)




    def create_menu(self):
        menubar = self.menuBar()

        # Crear el menú ESTANDARIZACION
        validaciones_menu = menubar.addMenu("Estandarización")
        validaciones_menu.addAction(QAction("Estandarizar MTJ para BD", self, triggered=self.estandarizarmtj))
        validaciones_menu.addAction(QAction("Estandarizar MTJ para BD BATCH", self, triggered=self.estandarizarmtjbatch))

        # Crear el menú VALIDACIÓN
        importar_menu = menubar.addMenu("Importar")
        importar_menu.addAction(QAction("Crear Tablas MTJ", self, triggered=self.crear_tablas))
        importar_menu.addAction(QAction("Cargar Datos a MTJ", self, triggered=self.cargar_datos))

        # Crear el menú EXPORTAR
        exportar_menu = menubar.addMenu("Exportar")
        exportar_menu.addAction(QAction("Exportar Formal/Informal", self, triggered=self.exportar_datos))
        exportar_menu.addAction(QAction("Exportar Formal/Informal con Naturaleza", self, triggered=self.exportar_datos_nat))
        exportar_menu.addAction(QAction("Exportar Unidades de Construcción", self, triggered=self.exportar_ucons))
        exportar_menu.addAction(QAction("Exportar consultas FILI para MTJ", self, triggered=self.exportar_consultas_fili))

        # Crear el menú ILI Tools
        ili_menu = menubar.addMenu("ILI Tools")
        gpkg_submenu = QMenu("GPKG Tools", self)
        gpkg_submenu.addAction(QAction("De GPKG A XTF", self, triggered=self.gpkgtoxt))
        gpkg_submenu.addAction(QAction("Importar XTF a GPKG", self, triggered=self.xtfgpkg))
        ili_menu.addMenu(gpkg_submenu)

        bd_submenu = QMenu("BD Tools", self)
        bd_submenu.addAction(QAction("Importar XTF a BD", self, triggered=self.xtfbd2))
        bd_submenu.addAction(QAction("Exportar BD a XTF", self, triggered=self.bd_a_xtf))
        ili_menu.addMenu(bd_submenu)

        borrado_submenu = QMenu("Depuración", self)
        borrado_submenu.addAction(QAction("Eliminar QR_DEFINITIVO", self, triggered=self.eliminado_qr))
        borrado_submenu.addAction(QAction("Mantener QR_DEFINITIVO", self, triggered=self.mantener_qr))
        borrado_submenu.addAction(QAction("Borrar todos los datos BD", self, triggered=self.eliminado_cascada))
        ili_menu.addMenu(borrado_submenu)

        # Crear el menú de limpieza
        limpiar_menu = menubar.addMenu("Limpiar Parámetros")
        limpiar_menu.addAction(QAction("Limpiar Parámetros", self, triggered=self.resetear_aplicacion))




    def agregar_ruta_psql(self):
        posibles_rutas = [
            Path("C:/Program Files/PostgreSQL/"),
            Path("C:/Program Files (x86)/PostgreSQL/")
        ]
        versiones = ["9.6", "10", "11", "12", "13", "14", "15", "16"]

        for base_path in posibles_rutas:
            for version in versiones:
                psql_path = base_path / version / "bin"
                if psql_path.exists() and (psql_path / "psql.exe").exists():
                    os.environ["PATH"] += os.pathsep + str(psql_path)
                    print(f"Ruta de psql añadida: {psql_path}")
                    return
        
        # Mostrar cuadro de error utilizando QMessageBox
        QMessageBox.critical(None, "Error", "No se encontró la instalación de psql en las ubicaciones comunes.")




    def gpkgtoxt(self):
        try:
            # Reemplaza 'abrir_ventana_exportacion' con la lógica necesaria para QGIS
            abrir_ventana_exportacion(self.iface)  # O cualquier lógica específica para la ventana de exportación
        except Exception as e:
            QMessageBox.critical(None, "Error durante la exportación", f"Error durante la exportación: {e}")

    def xtfgpkg(self):
        try:
            # Reemplaza 'abrir_ventana_importacion' con la lógica necesaria para QGIS
            abrir_ventana_importacion(self.iface)  # O cualquier lógica específica para la ventana de importación
        except Exception as e:
            QMessageBox.critical(None, "Error durante la exportación", f"Error durante la exportación: {e}")



    def xtfbd2(self):
        try:
            # Reemplaza 'abrir_ventana_importacionbd' con la lógica necesaria para QGIS
            abrir_ventana_importacionbd(self.iface, self.conexion, self.esquema_menu, self.entry_password.text())
        except Exception as e:
            QMessageBox.critical(None, "Error durante la exportación", f"Error durante la exportación: {e}")




    def bd_a_xtf(self):
        try:
            # Reemplaza 'abrir_ventana_exportacion_bd_2' con la lógica necesaria para QGIS
            abrir_ventana_exportacion_bd_2(self.iface, self.conexion, self.esquema_menu, self.entry_password.text())
        except Exception as e:
            QMessageBox.critical(None, "Error durante la exportación", f"Error durante la exportación: {e}")


    def crear_base_datos(self):
        # Usar QInputDialog para obtener el nombre de la base de datos
        nombre_base_datos, ok = QInputDialog.getText(self.iface.mainWindow(), "Crear Base de Datos", "Ingrese el nombre de la nueva base de datos FILI:")
        
        if ok and nombre_base_datos:
            try:
                # Conectar al servidor para crear la base de datos
                nueva_conexion = psycopg2.connect(
                    host=self.entry_host.text(),
                    port=self.entry_port.text(),
                    user=self.entry_usuario.text(),
                    password=self.entry_password.text()
                )
                nueva_conexion.autocommit = True

                cursor = nueva_conexion.cursor()

                # Verificar si la base de datos ya existe
                cursor.execute(f"SELECT 1 FROM pg_database WHERE datname = '{nombre_base_datos}';")
                if cursor.fetchone():
                    QMessageBox.warning(self.iface.mainWindow(), "Advertencia", f"La base de datos '{nombre_base_datos}' ya existe.")
                    return

                # Crear la nueva base de datos
                cursor.execute(f"CREATE DATABASE {nombre_base_datos};")
                nueva_conexion.close()

                # Conectar a la nueva base de datos para instalar las extensiones
                nueva_conexion_bd = psycopg2.connect(
                    host=self.entry_host.text(),
                    port=self.entry_port.text(),
                    database=nombre_base_datos,
                    user=self.entry_usuario.text(),
                    password=self.entry_password.text()
                )
                nueva_conexion_bd.autocommit = True

                cursor = nueva_conexion_bd.cursor()

                # Instalar extensiones PostGIS y uuid-ossp
                cursor.execute("CREATE EXTENSION IF NOT EXISTS postgis;")
                cursor.execute("CREATE EXTENSION IF NOT EXISTS \"uuid-ossp\";")

                # Insertar el sistema de coordenadas EPSG:9377 si no existe
                cursor.execute("""
                    INSERT INTO spatial_ref_sys (srid, auth_name, auth_srid, proj4text, srtext) 
                    VALUES (
                        9377, 
                        'EPSG', 
                        9377, 
                        '+proj=tmerc +lat_0=4.0 +lon_0=-73.0 +k=0.9992 +x_0=5000000 +y_0=2000000 +ellps=GRS80 +towgs84=0,0,0,0,0,0,0 +units=m +no_defs',
                        'PROJCRS["MAGNA-SIRGAS / Origen-Nacional", BASEGEOGCRS["MAGNA-SIRGAS", DATUM["Marco Geocentrico Nacional de Referencia", ELLIPSOID["GRS 1980",6378137,298.257222101, LENGTHUNIT["metre",1]]], PRIMEM["Greenwich",0, ANGLEUNIT["degree",0.0174532925199433]], ID["EPSG",4686]], CONVERSION["Colombia Transverse Mercator", METHOD["Transverse Mercator", ID["EPSG",9807]], PARAMETER["Latitude of natural origin",4, ANGLEUNIT["degree",0.0174532925199433], ID["EPSG",8801]], PARAMETER["Longitude of natural origin",-73, ANGLEUNIT["degree",0.0174532925199433], ID["EPSG",8802]], PARAMETER["Scale factor at natural origin",0.9992, SCALEUNIT["unity",1], ID["EPSG",8805]], PARAMETER["False easting",5000000, LENGTHUNIT["metre",1], ID["EPSG",8806]], PARAMETER["False northing",2000000, LENGTHUNIT["metre",1], ID["EPSG",8807]]], CS[Cartesian,2], AXIS["northing (N)",north, ORDER[1], LENGTHUNIT["metre",1]], AXIS["easting (E)",east, ORDER[2], LENGTHUNIT["metre",1]], USAGE[ SCOPE["unknown"], AREA["Colombia"], BBOX[-4.23,-84.77,15.51,-66.87]], ID["EPSG",9377]]'
                    ) 
                    ON CONFLICT (srid) DO NOTHING;
                """)

                nueva_conexion_bd.close()

                # Mostrar mensaje de éxito
                QMessageBox.information(self.iface.mainWindow(), "Éxito", f"La base de datos '{nombre_base_datos}' fue creada exitosamente, y las extensiones PostGIS y uuid-ossp fueron añadidas.")

                # Actualizar el combobox de bases de datos con la nueva base de datos creada
                self.conectar_y_obtener_esquemas()

            except psycopg2.Error as e:
                # Mostrar mensaje de error
                QMessageBox.critical(self.iface.mainWindow(), "Error", f"No se pudo crear la base de datos: {e}")









    def crear_esquema(self):
        # Crear un diálogo para ingresar el esquema
        ventana_esquema = QDialog(self.iface.mainWindow())
        ventana_esquema.setWindowTitle("Crear Esquema FILI (Formulario de Captura)")
        ventana_esquema.setFixedSize(420, 185)

        # Crear el layout principal
        layout = QVBoxLayout()

        # Etiqueta para ingresar el nombre del esquema
        etiqueta_nombre = QLabel("Ingrese el nombre del nuevo esquema FILI (Formulario de Captura ANT):")
        layout.addWidget(etiqueta_nombre)

        # Campo de texto para ingresar el nombre del esquema
        esquema_nombre_var = QLineEdit()
        layout.addWidget(esquema_nombre_var)

        # Checkbox para "Palmira"
        modelo_seleccionado = QCheckBox("Palmira (Solo activar para este mpio)")
        layout.addWidget(modelo_seleccionado)

        # Layout para los botones
        layout_botones = QHBoxLayout()

        # Botón para crear el esquema
        btn_crear = QPushButton("Crear esquema")
        layout_botones.addWidget(btn_crear)

        # Botón para cancelar
        btn_cancelar = QPushButton("Cancelar")
        layout_botones.addWidget(btn_cancelar)

        layout.addLayout(layout_botones)
        ventana_esquema.setLayout(layout)

        # Definir la función para crear el esquema
        def crear_esquema_y_modelo():
            nombre_esquema = esquema_nombre_var.text()
            if not nombre_esquema:
                QMessageBox.critical(self.iface.mainWindow(), "Error", "Debe ingresar un nombre para el esquema.")
                return

            # Selección de carpeta según el estado del checkbox
            if modelo_seleccionado.isChecked():
                modelos_ili_path = self.obtener_ruta_modelos_ili("Palmira")
            else:
                modelos_ili_path = self.obtener_ruta_modelos_ili("Demás municipios")

            modelo_ili_path = self.seleccionar_archivo_ili(modelos_ili_path)

            try:
                cursor = self.conexion.cursor()
                cursor.execute(f"CREATE SCHEMA IF NOT EXISTS {nombre_esquema};")
                self.conexion.commit()

                self.generar_modelo_fisico(nombre_esquema, modelo_ili_path, modelos_ili_path)

                # Volver a cargar los esquemas
                self.obtener_esquemas()
                ventana_esquema.accept()  # Cerrar la ventana de creación de esquema
            except Exception as e:
                QMessageBox.critical(self.iface.mainWindow(), "Error", f"No se pudo crear el esquema: {e}")

        # Asignar las funciones a los botones
        btn_crear.clicked.connect(crear_esquema_y_modelo)
        btn_cancelar.clicked.connect(ventana_esquema.reject)

        # Mostrar la ventana
        ventana_esquema.exec_()



    def obtener_ruta_modelos_ili(self, seleccion):
        """
        Obtiene la ruta a la carpeta que contiene los modelos .ili dentro de la estructura del proyecto.
        """
        current_dir = os.path.dirname(os.path.abspath(__file__))
        if seleccion == "Palmira":
            modelos_ili_path = os.path.join(current_dir, "codigos", "modelos2")
        else:
            modelos_ili_path = os.path.join(current_dir, "codigos", "modelos")
        
        if not os.path.exists(modelos_ili_path):
            raise FileNotFoundError(f"No se encontró la carpeta de modelos .ili en la ruta esperada: {modelos_ili_path}")
        
        return modelos_ili_path

    def seleccionar_archivo_ili(self, modelos_ili_path):
        """
        Lógica para seleccionar un archivo ILI dentro de la carpeta especificada.
        En este ejemplo, se seleccionará el primer archivo .ili encontrado.
        """
        for file_name in os.listdir(modelos_ili_path):
            if file_name.endswith(".ili"):
                return os.path.join(modelos_ili_path, file_name)
        raise FileNotFoundError(f"No se encontró ningún archivo .ili en la carpeta: {modelos_ili_path}")



    def generar_modelo_fisico(self, esquema, modelo_ili_path, modelos_ili_path):
        try:
            ili2pg_path = self.obtener_ruta_ili2pg()
            modelos = self.obtener_modelos_desde_ili(modelos_ili_path)

            comando = [
                "java",
                "-jar", ili2pg_path,
                "--schemaimport",
                "--dbhost", self.entry_host.text(),
                "--dbport", self.entry_port.text(),
                "--dbusr", self.entry_usuario.text(),
                "--dbpwd", self.entry_password.text(),
                "--dbdatabase", self.db_var.text(),
                "--dbschema", esquema,
                "--coalesceCatalogueRef",  # estructuras
                "--createNumChecks",  # creacion de restrucciones
                "--createUnique",  # creacion de restrucciones
                "--createFk",  # creacion de restrucciones
                "--createFkIdx",  # creacion de indices
                "--coalesceMultiSurface",  # estructuras
                "--coalesceMultiLine",  # estructuras
                "--beautifyEnumDispName",  # enumeracion
                "--createGeomIdx",  # creacion de indices
                "--createMetaInfo",  # metainformacion adicional
                "--createEnumTabsWithId",  # enumeracion
                "--smart2Inheritance",  # herencia
                "--strokeArcs",  # geometria
                "--defaultSrsAuth", "EPSG",  # geometria
                "--defaultSrsCode", "9377",  # geometria
                "--models", modelos,
                "--setupPgExt",  # ajustes
                modelo_ili_path
            ]

            resultado = subprocess.run(comando, capture_output=True, text=True)

            if resultado.returncode == 0:
                QMessageBox.information(None, "Éxito", f"La generación del modelo físico para FILI fue exitosa en el esquema '{esquema}'.")
            else:
                QMessageBox.critical(None, "Error", f"Error al generar el modelo físico ILI: {resultado.stderr}")

        except Exception as e:
            QMessageBox.critical(None, "Error", f"Error al intentar generar el modelo físico ILI: {e}")

    def obtener_ruta_ili2pg(self):
        """
        Obtiene la ruta al archivo ili2pg-5.1.0.jar dentro de la estructura del proyecto.
        """
        current_dir = os.path.dirname(os.path.abspath(__file__))
        ili2pg_path = os.path.join(current_dir, "codigos", "ilidb", "ili2pg-5.1.0.jar")
        
        if not os.path.exists(ili2pg_path):
            raise FileNotFoundError(f"No se encontró ili2pg en la ruta esperada: {ili2pg_path}")
        
        return ili2pg_path

    def obtener_modelos_desde_ili(self, modelos_ili_path):
        """
        Obtiene la lista de modelos a partir de los archivos .ili en la carpeta especificada.
        
        :param modelos_ili_path: Ruta a la carpeta que contiene los archivos .ili.
        :return: Una cadena con los nombres de los modelos separados por punto y coma.
        """
        modelos = []
        for file_name in os.listdir(modelos_ili_path):
            if file_name.endswith(".ili"):
                modelo = os.path.splitext(file_name)[0]
                modelos.append(modelo)
        
        return ";".join(modelos)





 ############################# ELIMINACION EN CASCADA ###############################################



    def eliminado_cascada(self):
        """
        Ejecuta la operación de eliminado en cascada.
        Crea la función de eliminación, cuenta los registros a eliminar, confirma la eliminación y la ejecuta.
        """
        try:
            # Obtener la conexión y validar que esté activa
            conexion = self.conectar_a_bd()
            if not conexion:
                QMessageBox.critical(None, "Error", "No se pudo conectar a la base de datos.")
                return

            esquema = self.esquema_menu.currentText()  # Obtener el esquema seleccionado
            if not esquema:
                QMessageBox.critical(None, "Error", "Debe seleccionar un esquema.")
                return

            # Crear la función en el esquema seleccionado
            self.crear_funcion_eliminar(esquema)

            # Contar los registros de las tablas antes de eliminarlos
            totales = self.contar_registros(esquema)

            if totales is None:
                return

            # Mostrar el mensaje de confirmación
            if self.confirmar_eliminacion(totales):
                # Si el usuario confirma, ejecutar la eliminación
                self.ejecutar_eliminacion(esquema)

        except Exception as e:
            QMessageBox.critical(None, "Error durante la eliminación", f"Error: {e}")



    def crear_funcion_eliminar(self, esquema):
        """
        Crea la función eliminar_todos_los_registros en el esquema seleccionado.
        """
        try:
            cursor = self.conexion.cursor()

            # Crear la función eliminar_todos_los_registros en el esquema seleccionado
            funcion_sql = f"""
            CREATE OR REPLACE FUNCTION {esquema}.eliminar_todos_los_registros()
            RETURNS VOID AS $$
            BEGIN
                -- Establecer el esquema predeterminado
                SET search_path TO {esquema};

                -- Eliminar las asociaciones en cca_extdireccion
                DELETE FROM cca_extdireccion;

                -- Eliminar las asociaciones relacionadas con cca_predio
                DELETE FROM cca_unidadconstruccion
                WHERE caracteristicasunidadconstruccion IN (
                    SELECT t_id
                    FROM cca_caracteristicasunidadconstruccion
                );

                DELETE FROM cca_caracteristicasunidadconstruccion;

                DELETE FROM cca_fuenteadministrativa;

                DELETE FROM cca_interesado;

                DELETE FROM cca_derecho;

                -- Eliminar registros de cca_terreno
                DELETE FROM cca_terreno;

                -- Eliminar registros de cca_predio;
                DELETE FROM cca_predio;

                -- Mensaje de éxito
                RAISE NOTICE 'Eliminación de todos los registros exitosa';
            END;
            $$ LANGUAGE plpgsql;
            """
            cursor.execute(funcion_sql)
            self.conexion.commit()
            cursor.close()

            QMessageBox.information(None, "Éxito", f"Función eliminar_todos_los_registros creada en el esquema {esquema}.")

        except psycopg2.Error as e:
            QMessageBox.critical(None, "Error", f"No se pudo crear la función de eliminación: {e}")


    def contar_registros(self, esquema):
        """
        Cuenta los registros de cada tabla antes de eliminarlos.
        """
        tablas = ['cca_extdireccion', 'cca_unidadconstruccion', 'cca_caracteristicasunidadconstruccion',
                'cca_fuenteadministrativa', 'cca_interesado', 'cca_derecho', 'cca_terreno', 'cca_predio']
        totales = {}
        try:
            cursor = self.conexion.cursor()
            for tabla in tablas:
                cursor.execute(f"SELECT COUNT(*) FROM {esquema}.{tabla};")
                total = cursor.fetchone()[0]
                totales[tabla] = total
            cursor.close()

        except psycopg2.Error as e:
            QMessageBox.critical(None, "Error", f"No se pudo contar los registros: {e}")
            return None

        return totales


    def confirmar_eliminacion(self, totales):
        """
        Muestra un mensaje de confirmación antes de proceder con la eliminación de los registros.
        """
        mensaje = "Se eliminarán los siguientes registros:\n\n"
        for tabla, total in totales.items():
            mensaje += f"- {total} registros de la tabla {tabla}\n"
        mensaje += "\n¿Desea continuar?"

        respuesta = QMessageBox.question(None, "Confirmar eliminación", mensaje, QMessageBox.Yes | QMessageBox.No)
        return respuesta == QMessageBox.Yes


    def ejecutar_eliminacion(self, esquema):
        """
        Ejecuta la función eliminar_todos_los_registros en el esquema seleccionado.
        """
        try:
            cursor = self.conexion.cursor()
            cursor.execute(f"SELECT {esquema}.eliminar_todos_los_registros();")
            self.conexion.commit()
            cursor.close()
            QMessageBox.information(None, "Éxito", "Los registros han sido eliminados exitosamente.")
        except psycopg2.Error as e:
            QMessageBox.critical(None, "Error", f"No se pudo ejecutar la eliminación: {e}")



 ############################# ELIMINACION POR QR###############################################
   

    def eliminado_qr(self):
        """
        Abre una ventana para que el usuario ingrese los QR a eliminar y ejecuta la eliminación.
        """
        try:
            # Conectar a la base de datos
            conexion = self.conectar_a_bd()
            if not conexion:
                QMessageBox.critical(None, "Error", "No se pudo conectar a la base de datos.")
                return

            esquema = self.esquema_menu.currentText()  # Obtener el esquema seleccionado
            if not esquema:
                QMessageBox.critical(None, "Error", "Debe seleccionar un esquema.")
                return

            # Crear la función eliminar_registros_qr si no existe
            self.crear_funcion_eliminar_qr(conexion, esquema)

            # Abrir ventana para ingresar los QR
            qr_input, ok = QInputDialog.getText(None, "Eliminar QR_DEFINITIVO", "Ingrese los QR a eliminar, separados por comas:")
            if not ok or not qr_input:
                QMessageBox.information(None, "Cancelado", "No se ingresaron QR para eliminar.")
                return

            # Formatear la entrada del usuario al formato requerido por la función
            qr_list = [qr.strip() for qr in qr_input.split(',') if qr.strip()]
            if not qr_list:
                QMessageBox.critical(None, "Error", "No se ingresaron QR válidos.")
                return

            # Confirmar eliminación
            qr_confirm = ', '.join(qr_list)
            confirm = QMessageBox.question(None, "Confirmación de eliminación", f"¿Desea eliminar los siguientes QR?\n{qr_confirm}", 
                                        QMessageBox.Yes | QMessageBox.No)

            if confirm == QMessageBox.Yes:
                # Ejecutar la función de eliminación
                cursor = conexion.cursor()
                # Preparar la lista para pasársela a la función SQL
                cursor.execute(f"SELECT {esquema}.eliminar_registros_qr(%s);", (qr_list,))
                conexion.commit()

                QMessageBox.information(None, "Éxito", f"Se han eliminado los QR: {qr_confirm}")
            else:
                QMessageBox.information(None, "Cancelado", "Eliminación cancelada.")
        except Exception as e:
            QMessageBox.critical(None, "Error durante la eliminación", f"Error: {e}")

    def crear_funcion_eliminar_qr(self, conexion, esquema):
        """
        Crea la función eliminar_registros_qr si no existe en el esquema dado.
        """
        try:
            cursor = conexion.cursor()

            # Crear la función en la base de datos si no existe
            cursor.execute(f"""
            CREATE OR REPLACE FUNCTION {esquema}.eliminar_registros_qr(qr_list TEXT[])
            RETURNS VOID AS $$
            DECLARE
            BEGIN
                SET search_path TO {esquema};

                -- Eliminar asociaciones y registros
                DELETE FROM cca_extdireccion
                WHERE cca_predio_direccion IN (
                    SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                );

                DELETE FROM cca_unidadconstruccion
                WHERE caracteristicasunidadconstruccion IN (
                    SELECT t_id FROM cca_caracteristicasunidadconstruccion WHERE predio IN (
                        SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                    )
                );

                DELETE FROM cca_caracteristicasunidadconstruccion
                WHERE predio IN (
                    SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                );

                DELETE FROM cca_fuenteadministrativa
                WHERE derecho IN (
                    SELECT t_id FROM cca_derecho WHERE predio IN (
                        SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                    )
                );

                DELETE FROM cca_interesado
                WHERE derecho IN (
                    SELECT t_id FROM cca_derecho WHERE predio IN (
                        SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                    )
                );

                DELETE FROM cca_derecho
                WHERE predio IN (
                    SELECT t_id FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                );

                DELETE FROM cca_terreno
                WHERE t_id IN (
                    SELECT terreno FROM cca_predio WHERE qr_operacion_definitivo = ANY(qr_list)
                );

                DELETE FROM cca_predio
                WHERE qr_operacion_definitivo = ANY(qr_list);

                RAISE NOTICE 'Eliminación de % exitosa', ARRAY_TO_STRING(qr_list, ', ');
            END;
            $$ LANGUAGE plpgsql;
            """)
            conexion.commit()
        except Exception as e:
            QMessageBox.critical(None, "Error", f"No se pudo crear la función eliminar_registros_qr: {e}")


 #############################  MANTENER REGISTROSS ################################################################

    def mantener_qr(self):
        """
        Abre una ventana para que el usuario ingrese los QR a mantener y ejecuta el proceso de eliminación del resto.
        """
        try:
            # Conectar a la base de datos
            conexion = self.conectar_a_bd()
            if not conexion:
                QMessageBox.critical(None, "Error", "No se pudo conectar a la base de datos.")
                return

            esquema = self.esquema_menu.currentText()  # Obtener el esquema seleccionado
            if not esquema:
                QMessageBox.critical(None, "Error", "Debe seleccionar un esquema.")
                return

            # Crear la función mantener_registros si no existe
            self.crear_funcion_mantener_qr(conexion, esquema)

            # Abrir ventana para ingresar los QR
            qr_input, ok = QInputDialog.getText(None, "Mantener QR_DEFINITIVO", "Ingrese los QR a mantener, separados por comas:")
            if not ok or not qr_input:
                QMessageBox.information(None, "Cancelado", "No se ingresaron QR para mantener.")
                return

            # Formatear la entrada del usuario al formato requerido por la función
            qr_list = [qr.strip() for qr in qr_input.split(',') if qr.strip()]
            if not qr_list:
                QMessageBox.critical(None, "Error", "No se ingresaron QR válidos.")
                return

            # Confirmar proceso
            qr_confirm = ', '.join(qr_list)
            confirm = QMessageBox.question(None, "Confirmación de mantenimiento", f"¿Desea mantener los siguientes QR y eliminar los demás?\n{qr_confirm}",
                                        QMessageBox.Yes | QMessageBox.No)

            if confirm == QMessageBox.Yes:
                # Ejecutar la función de mantenimiento
                cursor = conexion.cursor()

                # Preparar la lista como un array SQL (PostgreSQL necesita el array en formato específico)
                qr_array = '{' + ','.join(qr_list) + '}'

                # Ejecutar la función mantener_registros con el array
                cursor.execute(f"SELECT {esquema}.mantener_registros(%s);", (qr_array,))
                conexion.commit()

                QMessageBox.information(None, "Éxito", f"Se han mantenido los QR: {qr_confirm}, los demás han sido eliminados.")
            else:
                QMessageBox.information(None, "Cancelado", "El proceso de mantenimiento ha sido cancelado.")
        except Exception as e:
            QMessageBox.critical(None, "Error durante el mantenimiento", f"Error: {e}")


    def crear_funcion_mantener_qr(self, conexion, esquema):
        """
        Crea la función mantener_registros si no existe en el esquema dado.
        """
        try:
            cursor = conexion.cursor()

            # Crear la función en la base de datos si no existe
            cursor.execute(f"""
            CREATE OR REPLACE FUNCTION {esquema}.mantener_registros(qr_list TEXT[])
            RETURNS VOID AS $$
            DECLARE
            BEGIN
                SET search_path TO {esquema};

                -- Eliminar las asociaciones en cca_extdireccion
                DELETE FROM cca_extdireccion
                WHERE cca_predio_direccion IN (
                    SELECT t_id FROM cca_predio
                    WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                );

                -- Eliminar las asociaciones relacionadas con una lista de qr_operacion_definitivo en cca_predio
                DELETE FROM cca_unidadconstruccion
                WHERE caracteristicasunidadconstruccion IN (
                    SELECT t_id FROM cca_caracteristicasunidadconstruccion
                    WHERE predio IN (
                        SELECT t_id FROM cca_predio
                        WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                    )
                );

                DELETE FROM cca_caracteristicasunidadconstruccion
                WHERE predio IN (
                    SELECT t_id FROM cca_predio
                    WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                );

                DELETE FROM cca_fuenteadministrativa
                WHERE derecho IN (
                    SELECT t_id FROM cca_derecho
                    WHERE predio IN (
                        SELECT t_id FROM cca_predio
                        WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                    )
                );

                DELETE FROM cca_interesado
                WHERE derecho IN (
                    SELECT t_id FROM cca_derecho
                    WHERE predio IN (
                        SELECT t_id FROM cca_predio
                        WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                    )
                );

                DELETE FROM cca_derecho
                WHERE predio IN (
                    SELECT t_id FROM cca_predio
                    WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                );

                -- Eliminar registros de cca_terreno
                DELETE FROM cca_terreno
                WHERE t_id IN (
                    SELECT terreno
                    FROM cca_predio
                    WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1))
                );

                -- Eliminar registros de cca_predio
                DELETE FROM cca_predio
                WHERE qr_operacion_definitivo NOT IN (SELECT unnest($1));
                
                -- Mensaje de éxito
                RAISE NOTICE 'Eliminación de registros excluyendo % exitosa', ARRAY_TO_STRING($1, ', ');

            END;
            $$ LANGUAGE plpgsql;
            """)
            conexion.commit()
        except Exception as e:
            QMessageBox.critical(None, "Error", f"No se pudo crear la función mantener_registros: {e}")



  ############################# ################################################################


    def exportar_consultas_fili(self):
        esquema_seleccionado = self.esquema_menu.currentText()
        try:
            self.update_progress("Ejecutando consulta para exportación...")
            dataframes = ejecutar_consulta_fili(self.conexion, esquema_seleccionado)

            self.update_progress("Seleccionando ruta de guardado para el archivo Excel...")
            # Formatear el nombre del archivo para incluir el esquema seleccionado
            file_path, _ = QFileDialog.getSaveFileName(None, "Guardar archivo", f"CONSULTAS_FILI_PARA_MTJ_{esquema_seleccionado}.xlsx", "Excel files (*.xlsx)")
            
            if file_path:
                self.update_progress("Exportando consultas a Excel...")
                # Exportar los DataFrames a un archivo Excel
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    for sheet_name, df in dataframes.items():
                        self.update_progress(f"Exportando hoja: {sheet_name}...")
                        df.to_excel(writer, index=False, sheet_name=sheet_name)

                self.update_progress("Exportación completada.")

                # Crear un mensaje con un hipervínculo al archivo exportado
                msg_box = QMessageBox()
                msg_box.setIcon(QMessageBox.Information)
                msg_box.setWindowTitle("Exportación completada")
                msg_box.setText(f"Consulta exportada exitosamente. <a href='file:///{file_path}'>Abrir archivo</a>")
                msg_box.setStandardButtons(QMessageBox.Ok)

                # Habilitar los clics en el enlace
                msg_box.setTextInteractionFlags(Qt.TextBrowserInteraction)
                msg_box.exec_()

            else:
                self.update_progress("Exportación cancelada.")
        
        except Exception as e:
            self.update_progress(f"Error: {e}")
            QMessageBox.critical(None, "Error", str(e))




    def conectar_y_obtener_esquemas(self):
        host = self.entry_host.text()  # Obtener el texto de QLineEdit
        port = self.entry_port.text()  # Obtener el texto de QLineEdit
        usuario = self.entry_usuario.text()  # Obtener el texto de QLineEdit
        contraseña = self.entry_password.text()  # Obtener el texto de QLineEdit

        try:
            self.conexion = psycopg2.connect(
                host=host,
                port=port,
                user=usuario,
                password=contraseña
            )
            cursor = self.conexion.cursor()

            # Guardar la contraseña en QSettings después de conectarse con éxito
            settings = QSettings("ReportGeneratorApp", "BaseDatos")
            settings.setValue("ultima_contraseña", contraseña)


            # Listar todas las bases de datos
            cursor.execute("SELECT datname FROM pg_database WHERE datistemplate = false;")
            databases = cursor.fetchall()
            self.db_menu.clear()  # Limpiar el combobox de bases de datos
            self.db_menu.addItems([db[0] for db in databases])  # Añadir bases de datos al ComboBox

            # Asociar el evento de selección de base de datos al combobox
            self.db_menu.currentIndexChanged.connect(self.obtener_esquemas)
            #QMessageBox.information(self, "Éxito", "Conectado al servidor de base de datos. Seleccione una base de datos.")
        except psycopg2.OperationalError as e:
            QMessageBox.critical(self, "Error", f"Error al conectar a la base de datos: {e}")


    def obtener_esquemas(self, event=None):
        nombre_base_datos = self.db_menu.currentText()  # Obtener el valor seleccionado en el ComboBox
        cursor = self.conexion.cursor()

        try:
            # Cerrar la conexión existente y crear una nueva para la base de datos seleccionada
            self.conexion.close()
            self.conexion = psycopg2.connect(
                host=self.entry_host.text(),
                port=self.entry_port.text(),
                database=nombre_base_datos,
                user=self.entry_usuario.text(),
                password=self.entry_password.text()
            )
            cursor = self.conexion.cursor()

            # Ejecutar cualquier extensión que necesites crear
            self.crear_extensiones(cursor)



            # Listar todos los esquemas de la base de datos seleccionada
            cursor.execute("SELECT schema_name FROM information_schema.schemata;")
            esquemas = cursor.fetchall()

            # Limpiar y rellenar el combobox de esquemas
            self.esquema_menu.clear()  # Limpiar el combobox de esquemas
            if esquemas:
                self.esquema_menu.addItems([esquema[0] for esquema in esquemas])

            #QMessageBox.information(self, "Éxito", f"Esquemas cargados para la base de datos '{nombre_base_datos}'.")

        except psycopg2.OperationalError as e:
            QMessageBox.critical(self, "Error", f"Error al conectar a la base de datos: {e}")



    def conectar_a_bd(self):
        host = self.entry_host.text()  # Usar .text() en lugar de .get()
        port = self.entry_port.text()  # Usar .text() en lugar de .get()
        nombre_base_datos = self.db_menu.currentText()  # Usar .currentText() en lugar de .get()

        usuario = self.entry_usuario.text()  # Usar .text() en lugar de .get()
        contraseña = self.entry_password.text()  # Usar .text() en lugar de .get()

        try:
            conexion = psycopg2.connect(
                host=host,
                port=port,
                database=nombre_base_datos,
                user=usuario,
                password=contraseña
            )
            return conexion
        except psycopg2.OperationalError as e:
            QMessageBox.critical(self, "Error", f"Error al conectar a la base de datos: {e}")
            return None






    def crear_extensiones(self, cursor):
        """
        Crea las extensiones necesarias en la base de datos seleccionada.
        """
        try:
            # Verificar si la extensión 'postgis' está instalada
            cursor.execute("SELECT 1 FROM pg_extension WHERE extname = 'postgis';")
            postgis_exists = cursor.fetchone()

            if not postgis_exists:
                cursor.execute("CREATE EXTENSION IF NOT EXISTS postgis CASCADE;")
            
            # Verificar si la extensión 'postgis_topology' está instalada
            cursor.execute("SELECT 1 FROM pg_extension WHERE extname = 'postgis_topology';")
            postgis_topology_exists = cursor.fetchone()

            if not postgis_topology_exists:
                cursor.execute("CREATE EXTENSION IF NOT EXISTS postgis_topology CASCADE;")
            
            # Verificar si la extensión 'unaccent' está instalada
            cursor.execute("SELECT 1 FROM pg_extension WHERE extname = 'unaccent';")
            unaccent_exists = cursor.fetchone()

            if not unaccent_exists:
                cursor.execute("CREATE EXTENSION IF NOT EXISTS unaccent;")
            
            # Verificar si la extensión 'uuid-ossp' está instalada
            cursor.execute("SELECT 1 FROM pg_extension WHERE extname = 'uuid-ossp';")
            uuid_ossp_exists = cursor.fetchone()

            if not uuid_ossp_exists:
                cursor.execute("CREATE EXTENSION IF NOT EXISTS \"uuid-ossp\";")
            
            # Mostrar mensaje de éxito en consola o GUI
            print("Extensiones PostGIS, PostGIS Topology, Unaccent y UUID-OSSP instaladas correctamente.")
        
        except psycopg2.Error as e:
            # Mostrar mensaje de error en caso de fallo al instalar las extensiones
            QMessageBox.critical(self.iface.mainWindow(), "Error", f"Error al crear extensiones: {e}")




    def crear_tablas(self):
        # Verificar si hay una conexión
        if not self.conexion:
            QMessageBox.critical(None, "Error", "No hay conexión a la base de datos.")
            return

        try:
            cursor = self.conexion.cursor()
            consulta_crear_tablas = """
                CREATE SCHEMA IF NOT EXISTS mtj;

                CREATE TABLE IF NOT EXISTS mtj.matriz_mtj (
                    "id" bigint,
                    "uit" TEXT,
                    "local_id" TEXT,
                    "id_operacion" TEXT,
                    "fecha_asig" TEXT,
                    "profesional_tecnico" TEXT,
                    "fecha_analisi_tecnico" TEXT,
                    "estado_tecnico" TEXT,
                    "observaciones_tecnico" TEXT,
                    "CR_InteresadoTipo" TEXT,
                    "CR_DocumentoTipo" TEXT,
                    "CR_Interesado_all" TEXT,
                    "Documento_Identidad" TEXT,
                    "jefe_hogar" TEXT,
                    "victima_conflicto" TEXT,
                    "estado_civil" TEXT,
                    "inicio_tenencia" TEXT,
                    "reside_predio" TEXT,
                    "persona_distinta_reside" TEXT,
                    "quien" TEXT,
                    "explota_predio" TEXT,
                    "metodo_levan" TEXT,
                    "area_terreno_levantamiento" TEXT,
                    "nombre_predio" TEXT,
                    "destino_economico" TEXT,
                    "tiene_rezago" TEXT,
                    "num_rezago" TEXT,
                    "area_lpp_resgitral" TEXT,
                    "area_dif_catas_resgi" TEXT,
                    "tiene_construcciones" TEXT,
                    "cons_cantidad" TEXT,
                    "cons_area" TEXT,
                    "departamento" TEXT,
                    "municipio" TEXT,
                    "vereda" TEXT,
                    "numero_predial" TEXT,
                    "numero_predial_provisional" TEXT,
                    "propietario_catstral" TEXT,
                    "direccion_catastral" TEXT,
                    "area_catastral" TEXT,
                    "area_catastral_poligono" TEXT,
                    "area_cons_catastral" TEXT,
                    "destino_economico_catastral" TEXT,
                    "fecha_consulta_catastral" TEXT,
                    "tipo_novedad" TEXT,
                    "complemneto_tipo_novedad" TEXT,
                    "clasificacion_suelo" TEXT,
                    "uso_suelo_plan_parcial" TEXT,
                    "categ_suelo_rural" TEXT,
                    "categoria_proteccion" TEXT,
                    "desarrollo_restringido" TEXT,
                    "fecha_certificacion_uso_suelo" TEXT,
                    "cruza_determinantes" TEXT,
                    "cruza_desc_restricciones" TEXT,
                    "cruza_desc_restricciones_area_por" TEXT,
                    "cruza_desc_condicionantes" TEXT,
                    "cruza_desc_condicionantes_area_por" TEXT,
                    "cruza_area_habilitada" TEXT,
                    "concepto_catastral_general" TEXT,
                    "concepto_catastral_especifico" TEXT,
                    "req_acta_colindancia" TEXT,
                    "req_uso_suelo" TEXT,
                    "res_certificado_riesgos" TEXT,
                    "profesional_juridico" TEXT,
                    "fecha_analisi_juridico" TEXT,
                    "estado_juridico" TEXT,
                    "observaciones_juridico" TEXT,
                    "formal_informal" TEXT,
                    "naturaleza_predio" TEXT,
                    "tipo_predio" TEXT,
                    "tipo_derecho" TEXT,
                    "condicion_predio" TEXT,
                    "relaciona_fmi" TEXT,
                    "Codigo_Orip" TEXT,
                    "Matricula_Inmobiliaria" TEXT,
                    "ExtReferenciaRegistralSistemaAntiguo" TEXT,
                    "fmi_propietario" TEXT,
                    "fmi_ref_catastral" TEXT,
                    "fmi_area" TEXT,
                    "fmi_direccion" TEXT,
                    "fmi_tipo_predio" TEXT,
                    "fmi_matriz" TEXT,
                    "fmi_derivados" TEXT,
                    "tipo_fuente_adm" TEXT,
                    "fmi_ente_emisor" TEXT,
                    "numero_fuente_adm" TEXT,
                    "fmi_fecha_documento_fuente" TEXT,
                    "disponibilidad_fuente_adm" TEXT,
                    "documento_apertura_fmi" TEXT,
                    "fecha_doc_acto_apertura" TEXT,
                    "fmi_medidas_cautelares" TEXT,
                    "fmi_salvedades_correc" TEXT,
                    "fmi_complementa" TEXT,
                    "fmi_cabida_lindero" TEXT,
                    "fmi_titulo_originario" TEXT,
                    "fmi_estado" TEXT,
                    "tipo_novedad_fmi" TEXT,
                    "estado_rtdaf" TEXT,
                    "rupta" TEXT,
                    "predio_cruza_pretension" TEXT,
                    "tipo_pretension" TEXT,
                    "medida_2333" TEXT,
                    "fadc" TEXT,
                    "req_ospr" TEXT,
                    "concepto_juridico" TEXT,
                    "tipologia_titulacion" TEXT,
                    "ruta_atencion" TEXT,
                    "fmi_procedimientos_catatrales" TEXT,
                    "material_cimple" TEXT,
                    "doc_faltantes" TEXT,
                    "habilitado_reso" TEXT,
                    "reso" TEXT,
                    "profesional_agro" TEXT,
                    "fecha_analisi_agro" TEXT,
                    "estado_agro" TEXT,
                    "observaciones_agro" TEXT,
                    "agro_metodo_analisis" TEXT,
                    "agro_ufh" TEXT,
                    "agro_cober_agro" TEXT,
                    "agrp_fuente_info" TEXT,
                    "agro_aptitud" TEXT,
                    "agro_cultivos_ilicitos" TEXT,
                    "agro_recomendaciones" TEXT,
                    "agro_concepto" TEXT
                );

                CREATE TABLE IF NOT EXISTS mtj.interesado_mtj (
                    "qr_derivado" TEXT,
                    "interesado_tipo" TEXT,
                    "documento_tipo" TEXT,
                    "primer_nombre" TEXT,
                    "segundo_nombre" TEXT,
                    "primer_apellido" TEXT,
                    "segundo_apellido" TEXT,
                    "interesado_concat" TEXT,
                    "documento_numero" TEXT,
                    "jefe_hogar" TEXT,
                    "victima_conflicto" TEXT,
                    "estado_civil" TEXT,
                    "fecha_inicio_tenencia" TEXT,
                    "reside_predio_interesado" TEXT,
                    "reside_distinta_interesado" TEXT,
                    "quien" TEXT,
                    "explota_directamente" TEXT,
                    "numero_formulario_reso" TEXT
                );
            """
            cursor.execute(consulta_crear_tablas)
            self.conexion.commit()
            cursor.close()

            # Mensaje de éxito corregido
            QMessageBox.information(None, "Éxito", "Las tablas se han creado correctamente.")
        except psycopg2.Error as e:
            # Mensaje de error corregido
            QMessageBox.critical(None, "Error", f"Error al crear las tablas: {e}")









    def cargar_datos(self):
        if not getattr(self, 'conexion', None):
            QMessageBox.critical(None, "Error", "Por favor, conectarse a la base de datos primero.")
            return

        def detect_encoding(file_path):
            with open(file_path, 'rb') as f:
                result = chardet.detect(f.read())
            return result['encoding']

        try:
            cursor = self.conexion.cursor()

            cursor.execute("SELECT schema_name FROM information_schema.schemata WHERE schema_name = 'mtj';")
            mtj_schema = cursor.fetchone()
            if not mtj_schema:
                QMessageBox.critical(None, "Error", "El esquema MTJ no existe.")
                return

            cursor.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'mtj' AND table_name = 'matriz_mtj';")
            matriz_mtj_table = cursor.fetchone()

            if not matriz_mtj_table:
                QMessageBox.critical(None, "Error", "Las tablas necesarias no existen en el esquema MTJ.")
                return

            self.update_progress("Seleccionando archivos CSV...")

            ruta_archivo_mtj, _ = QFileDialog.getOpenFileName(None, "Seleccione el archivo CSV MTJ Estandarizados", "", "Archivos CSV (*.csv)")

            if ruta_archivo_mtj:
                try:
                    self.update_progress("Detectando la codificación de los archivos CSV...")
                    encoding_mtj = detect_encoding(ruta_archivo_mtj)

                    self.update_progress("Cargando datos del archivo CSV MTJ a la base de datos...")
                    with open(ruta_archivo_mtj, 'r', encoding=encoding_mtj) as f:
                        cursor.copy_expert("COPY mtj.matriz_mtj FROM STDIN WITH CSV HEADER DELIMITER ';' ENCODING 'UTF8'", f)

                    self.conexion.commit()
                    self.update_progress("Datos cargados exitosamente.")
                    QMessageBox.information(None, "Éxito", "Datos cargados exitosamente.")
                except psycopg2.Error as e:
                    self.update_progress(f"Error al cargar los datos: {e}")
                    QMessageBox.critical(None, "Error", f"Error al cargar los datos: {e}")
                except FileNotFoundError as e:
                    self.update_progress(f"Error al abrir el archivo: {e}")
                    QMessageBox.critical(None, "Error", f"Error al abrir el archivo: {e}")
                finally:
                    if self.conexion:
                        self.conexion.close()
            else:
                QMessageBox.critical(None, "Error", "Debe seleccionar el archivo CSV.")
        except psycopg2.Error as e:
            QMessageBox.critical(None, "Error", f"Error al realizar la comprobación de esquema y tablas: {e}")





    def unir_interesados(self):
        if not getattr(self, 'conexion', None):
            QMessageBox.critical(None, "Error", "Por favor, conectarse a la base de datos primero.")
            return

        esquema_seleccionado = self.esquema_var.get()
        if not esquema_seleccionado:
            QMessageBox.critical(None, "Error", "Por favor, seleccionar un esquema primero.")
            return

        try:
            self.update_progress("Verificando registros existentes en la tabla 'mtj.cca_interesado2'...")

            cursor = self.conexion.cursor()
            cursor.execute("SELECT fuente, COUNT(*) FROM mtj.cca_interesado2 GROUP BY fuente")

            registros_por_fuente = cursor.fetchall()

            if registros_por_fuente:
                mensaje = "La tabla ya tiene registros:\n"
                for fuente, cantidad in registros_por_fuente:
                    mensaje += f"- {cantidad} registros en {fuente}\n"
                mensaje += "¿Desea continuar y potencialmente duplicar la información?"

                confirmacion = QMessageBox.question(None, "Confirmar", mensaje, QMessageBox.Yes | QMessageBox.No)
                if confirmacion == QMessageBox.No:
                    self.update_progress("Operación cancelada por el usuario.")
                    return

            self.update_progress("Ejecutando la inserción de datos en 'mtj.cca_interesado2'...")

            cursor.execute(f"""
                INSERT INTO mtj.cca_interesado2 (
                    qr_operacion_definitivo,
                    condicion_predio,
                    interesado_tipo,
                    tipo_documento,
                    documento_identidad,
                    primer_nombre,
                    segundo_nombre,
                    primer_apellido,
                    segundo_apellido,
                    nombre_completo,
                    fuente
                )
                SELECT 
                    pd.qr_operacion_definitivo,
                    cpt.ilicode as condicion_predio,
                    int.ilicode AS interesado_tipo,
                    indt.ilicode AS tipo_documento,
                    i.documento_identidad AS documento_identidad,
                    i.primer_nombre,
                    i.segundo_nombre,
                    i.primer_apellido,
                    i.segundo_apellido,
                    STRING_AGG(CONCAT(i.primer_nombre, ' ', i.segundo_nombre, ' ', i.primer_apellido, ' ', i.segundo_apellido), ';') AS nombre_completo, 
                    'BD' as fuente
                FROM 
                    {esquema_seleccionado}.cca_predio pd
                JOIN {esquema_seleccionado}.cca_condicionprediotipo cpt ON pd.condicion_predio=cpt.t_id
                JOIN {esquema_seleccionado}.cca_derecho dr ON pd.t_id = dr.predio
                JOIN {esquema_seleccionado}.cca_interesado i ON dr.t_id = i.derecho
                JOIN {esquema_seleccionado}.cca_interesadotipo int ON i.tipo = int.t_id
                JOIN {esquema_seleccionado}.cca_interesadodocumentotipo indt ON i.tipo_documento = indt.t_id
                GROUP BY 
                    pd.qr_operacion_definitivo, int.ilicode, indt.ilicode, i.documento_identidad, i.primer_nombre, i.segundo_nombre, i.primer_apellido, i.segundo_apellido, cpt.ilicode
                UNION 
                SELECT 
                    imtj.qr_derivado as qr_operacion_definitivo,
                    mtj1.condicion_predio as condicion_predio,
                    imtj.interesado_tipo AS interesado_tipo,
                    imtj.documento_tipo AS tipo_documento,
                    imtj.documento_numero AS documento_identidad,
                    imtj.primer_nombre, 
                    imtj.segundo_nombre,
                    imtj.primer_apellido, 
                    imtj.segundo_apellido,
                    STRING_AGG(CONCAT(imtj.primer_nombre, ' ', imtj.segundo_nombre, ' ',imtj.primer_apellido, ' ', imtj.segundo_apellido), ';') AS nombre_completo, 
                    'MTJ' as fuente
                FROM  
                    mtj.interesado_mtj as imtj
                LEFT JOIN 
                    mtj.matriz_mtj mtj1 ON imtj.qr_derivado = mtj1.id_operacion
                GROUP BY 
                    qr_derivado, interesado_tipo, documento_tipo, documento_numero, primer_nombre, segundo_nombre, primer_apellido, segundo_apellido, mtj1.condicion_predio
                ORDER BY 
                    documento_identidad;
            """)

            self.update_progress("Datos insertados correctamente en 'mtj.cca_interesado2'.")
            QMessageBox.information(None, "Éxito", "Consulta ejecutada y datos cargados en mtj.cca_interesado2.")
            self.conexion.commit()
        except psycopg2.Error as e:
            self.update_progress(f"Error al ejecutar la consulta: {e}")
            QMessageBox.critical(None, "Error", f"Error al ejecutar la consulta: {e}")




    def resetear_aplicacion(self):
        # Restablecer el host
        self.entry_host.clear()
        self.entry_host.setText("localhost")
        
        # Restablecer el puerto
        self.entry_port.clear()
        self.entry_port.setText("5432")
        
        # Limpiar la selección de base de datos
        self.db_menu.setCurrentIndex(-1)  # Limpiar la selección actual del combobox
        
        # Restablecer el usuario
        self.entry_usuario.clear()
        self.entry_usuario.setText("postgres")
        
        # Limpiar la contraseña
        self.entry_password.clear()
        
        # Limpiar el cuadro de texto del estado detallado del progreso
        self.status_text.clear()

        # Cerrar la conexión si existe y eliminar el atributo
        if hasattr(self, 'conexion') and self.conexion is not None:
            self.conexion.close()
            delattr(self, 'conexion')
        
        # Limpiar los menús de esquema y reporte
        self.esquema_menu.setCurrentIndex(-1)
        self.reporte_menu.setCurrentIndex(-1)

        # Limpiar el cuadro de estado
        self.update_progress("")



    def estandarizarmtj(self):
        try:
            # Llama a la función de estandarización
            ejecutar_estandarizacion(self.update_progress)

            # Mensaje de éxito: El primer argumento de QMessageBox es el "parent" (self en este caso)
            QMessageBox.information(self, "Estandarización completada", "Estandarización completada exitosamente.")
        
        except Exception as e:
            # Mensaje de error: El primer argumento debe ser el parent, y no un string
            QMessageBox.critical(self, "Error durante la estandarización", f"Ocurrió un error durante la estandarización: {e}")



    def estandarizarmtjbatch(self):
        try:
            # Llamar a la ventana de estandarización batch
            open_estandarizacion_window()  
            
        except Exception as e:
            # Si hay un error, mostrar un mensaje crítico
            QMessageBox.critical(None, "Error durante la estandarización", f"Ocurrió un error durante la estandarización: {e}")



    def exportar_datos(self):
        self.update_progress("Conectándose a la base de datos...")
        conexion = self.conectar_a_bd()
        
        if conexion:
            esquema = self.esquema_menu.currentText()  # Obtener el esquema seleccionado
            
            self.update_progress("Seleccionando carpeta de destino para la exportación...")
            output_folder = QFileDialog.getExistingDirectory(None, "Seleccione la carpeta de destino")  # Seleccionar carpeta
            
            if output_folder:  
                try:
                    self.update_progress("Exportando datos a GPKG...")
                    export_to_geopackage(conexion, output_folder, esquema)  # Exportar los datos a GPKG
                    self.update_progress("Exportación completada.")
                    QMessageBox.information(None, "Éxito", "Datos exportados correctamente a GPKG.")
                except Exception as e:
                    self.update_progress(f"Error al exportar datos: {e}")
                    QMessageBox.critical(None, "Error", f"Error al exportar datos: {e}")
            else:
                self.update_progress("No se seleccionó ninguna carpeta de destino.")
                QMessageBox.warning(None, "Advertencia", "No se seleccionó ninguna carpeta de destino.")
        else:
            self.update_progress("Error al conectarse a la base de datos.")
            QMessageBox.critical(None, "Error", "Error al conectarse a la base de datos.")




    def exportar_ucons(self):
        self.update_progress("Conectándose a la base de datos...")
        conexion = self.conectar_a_bd()
        
        if conexion:
            esquema = self.esquema_menu.currentText()  
            
            self.update_progress("Seleccionando carpeta de destino para la exportación...")
            output_folder = QFileDialog.getExistingDirectory(None, "Seleccione la carpeta de destino")  # Seleccionar carpeta
            
            if output_folder:  
                try:
                    self.update_progress("Exportando datos a GPKG...")
                    export_to_geopackage3(conexion, output_folder, esquema)  # Exportar los datos a GPKG
                    self.update_progress("Exportación completada.")
                    QMessageBox.information(None, "Éxito", "Datos exportados correctamente a GPKG.")
                except Exception as e:
                    self.update_progress(f"Error al exportar datos: {e}")
                    QMessageBox.critical(None, "Error", f"Error al exportar datos: {e}")
            else:
                self.update_progress("No se seleccionó ninguna carpeta de destino.")
                QMessageBox.warning(None, "Advertencia", "No se seleccionó ninguna carpeta de destino.")
        else:
            self.update_progress("Error al conectarse a la base de datos.")
            QMessageBox.critical(None, "Error", "Error al conectarse a la base de datos.")



    def exportar_datos_nat(self):
        self.update_progress("Conectándose a la base de datos...")
        conexion = self.conectar_a_bd()

        if conexion:
            esquema = self.esquema_menu.currentText()  # Obtener el esquema seleccionado

            try:
                cursor = conexion.cursor()
                self.update_progress("Verificando si el esquema 'mtj' existe...")
                # Verificar si el esquema 'mtj' existe
                cursor.execute("SELECT schema_name FROM information_schema.schemata WHERE schema_name = 'mtj'")
                esquema_existe = cursor.fetchone()

                if esquema_existe:
                    self.update_progress("Verificando si la tabla 'matriz_mtj' contiene registros...")
                    # Verificar si la tabla 'matriz_mtj' contiene registros
                    cursor.execute("SELECT COUNT(*) FROM mtj.matriz_mtj")
                    registros = cursor.fetchone()[0]

                    if registros > 0:
                        self.update_progress("Verificando valores diferentes en 'naturaleza_predio'...")
                        # Verificar si hay valores diferentes en naturaleza_predio, incluyendo nulos o vacíos
                        cursor.execute(f"""
                            SELECT COALESCE(mmtj.naturaleza_predio, 'SIN DILIGENCIAR') AS naturaleza_predio, COUNT(*) 
                            FROM {esquema}.cca_predio pd
                            JOIN mtj.matriz_mtj mmtj ON pd.qr_operacion_definitivo = mmtj.id_operacion
                            WHERE COALESCE(mmtj.naturaleza_predio, '') NOT IN ('PRIVADO', 'privado', 'Privado', 'PUBLICO', 'PÚBLICO', 'Público', 'Publico')
                            GROUP BY naturaleza_predio
                        """)
                        valores_diferentes = cursor.fetchall()

                        if valores_diferentes:
                            valores_texto = '\n'.join([f"{fila[0]}: {fila[1]}" for fila in valores_diferentes])
                            respuesta = QMessageBox.question(None, "Advertencia", 
                                f"HAY VALORES DIFERENTES A PUBLICO/PRIVADO EN NATURALEZAS DEL PREDIO DE LA MTJ, SE DEBE AJUSTAR:\n{valores_texto}\n\n¿Desea continuar?",
                                QMessageBox.Yes | QMessageBox.No)
                            
                            if respuesta == QMessageBox.No:
                                self.update_progress("Operación cancelada por el usuario.")
                                return  # Detener el proceso si el usuario selecciona "No"
                        else:
                            self.update_progress("No se encontraron valores diferentes en naturaleza_predio.")

                        # Proceder con la exportación si todo está correcto
                        self.update_progress("Seleccionando carpeta de destino para la exportación...")
                        output_folder = QFileDialog.getExistingDirectory(None, "Seleccione la carpeta de destino")
                        if output_folder:
                            try:
                                self.update_progress("Exportando datos a GPKG...")
                                export_to_geopackage2(conexion, output_folder, esquema)
                                self.update_progress("Exportación completada.")
                                QMessageBox.information(None, "Éxito", "Datos de formales/informales con sus Naturalezas exportadas correctamente a GPKG.")
                            except Exception as e:
                                self.update_progress(f"Error al exportar datos: {e}")
                                QMessageBox.critical(None, "Error", f"Error al exportar datos: {e}")
                        else:
                            self.update_progress("No se seleccionó ninguna carpeta de destino.")
                            QMessageBox.warning(None, "Advertencia", "No se seleccionó ninguna carpeta de destino.")
                    else:
                        self.update_progress("No se ha cargado información de la MTJ a las Base de Datos.")
                        QMessageBox.warning(None, "Advertencia", "No se ha cargado información de la MTJ a las Base de Datos")
                else:
                    self.update_progress("El esquema 'mtj' no existe en la base de datos.")
                    QMessageBox.critical(None, "Error", "El esquema 'mtj' no existe en la base de datos.")

            except Exception as e:
                self.update_progress(f"Error al verificar esquema o tabla: {e}")
                QMessageBox.critical(None, "Error", f"Error al verificar esquema o tabla: {e}")
        else:
            self.update_progress("No se pudo conectar a la base de datos.")
            QMessageBox.critical(None, "Error", "No se pudo conectar a la base de datos.")






    def verificar_esquema_mtj(self):
        try:
            cursor = self.conexion.cursor()
            
            # Verificar si el esquema 'mtj' existe
            cursor.execute("SELECT schema_name FROM information_schema.schemata WHERE schema_name = 'mtj'")
            esquema_existe = cursor.fetchone()
            
            if esquema_existe:
                # Verificar si la tabla 'matriz_mtj' contiene registros
                cursor.execute("SELECT COUNT(*) FROM mtj.matriz_mtj")
                registros = cursor.fetchone()[0]
                
                if registros > 0:
                    return True
                else:
                    QMessageBox.warning(None, "Advertencia", "No se ha cargado información de la MTJ a las Base de Datos")
                    return False
            else:
                QMessageBox.critical(None, "Error", "El esquema 'mtj' no existe en la base de datos.")
                return False
            
        except Exception as e:
            QMessageBox.critical(None, "Error", f"Error al verificar esquema o tabla: {e}")
            return False


    def ejecutar_consultas(self):
        self.update_progress("Iniciando generación de reporte...")

        # Obtener el esquema seleccionado del combobox
        esquema_seleccionado = self.esquema_menu.currentText()
        # Obtener el tipo de reporte seleccionado del combobox
        tipo_reporte = self.reporte_menu.currentText()

        if not esquema_seleccionado:
            QMessageBox.critical(None, "Error", "Debe seleccionar un esquema.")
            return
        if not tipo_reporte:
            QMessageBox.critical(None, "Error", "Debe seleccionar un tipo de reporte.")
            return

        # Verificar existencia del esquema y datos en 'mtj' si se selecciona tipo de reporte 'Unificado' o 'BD MTJ Observaciones'
        if tipo_reporte in ["Unificado", "BD MTJ Observaciones"]:
            if not self.verificar_esquema_mtj():
                return

        ruta_archivo = QFileDialog.getExistingDirectory(None, "Seleccione la ruta para guardar el archivo de reporte")
        if not ruta_archivo:
            QMessageBox.critical(None, "Error", "Debe seleccionar una ruta para guardar el archivo de reporte.")
            return

        configuraciones_sql = [
            f"SET search_path TO {esquema_seleccionado}, public;"
        ]

        # Obtener el timestamp actual
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

        # Definir variables para los casos individuales y unificados
        if tipo_reporte == "Observaciones BD":
            from .codigos.consultas_reporte_1_bd import consultas_reporte_1_bd, nombres_consultas_observaciones
            consultas = consultas_reporte_1_bd
            nombres_consultas = nombres_consultas_observaciones
            nombre_archivo_final = f"{ruta_archivo}/1. REPORTE_CONSISTENCIA LÓGICA_BD_{esquema_seleccionado}_{timestamp}.xlsx"
            nombre_archivo_basico = f"{ruta_archivo}/1. Reporte_simplificado_Consistencia_Lógica_BD_{esquema_seleccionado}_{timestamp}.xlsx"
        
        elif tipo_reporte == "BD MTJ Observaciones":
            from .codigos.consultas_reporte_2_mtj import consultas_reporte_2_mtj, nombres_consultas_observaciones_mtj
            consultas = [consulta.format(esquema=esquema_seleccionado) for consulta in consultas_reporte_2_mtj]
            nombres_consultas = nombres_consultas_observaciones_mtj
            nombre_archivo_final = f"{ruta_archivo}/2. REPORTE_BD_vs_MTJ_OBSERVACIONES_{esquema_seleccionado}_{timestamp}.xlsx"
            nombre_archivo_basico = f"{ruta_archivo}/2. Reporte_simplificado_BD_vs_MTJ__{esquema_seleccionado}_{timestamp}.xlsx"
        
        elif tipo_reporte == "Unificado":
            from .codigos.consultas_reporte_1_bd import consultas_reporte_1_bd, nombres_consultas_observaciones
            from .codigos.consultas_reporte_2_mtj import consultas_reporte_2_mtj, nombres_consultas_observaciones_mtj
            consultas_bd = consultas_reporte_1_bd
            nombres_consultas_bd = nombres_consultas_observaciones
            consultas_mtj = [consulta.format(esquema=esquema_seleccionado) for consulta in consultas_reporte_2_mtj]
            nombres_consultas_mtj = nombres_consultas_observaciones_mtj
            nombre_archivo_final = f"{ruta_archivo}/3. Reporte_Unificado_BD_MTJ_{timestamp}.xlsx"
            nombre_archivo_basico = f"{ruta_archivo}/3. Reporte_simplificado_Unificado_BD_MTJ_{timestamp}.xlsx"




        # Generar el archivo básico sin formato para visualización rápida


        # Comprobar si el reporte es consolidado o simplificado
        if self.simplificado_radio.isChecked():
            # Si el reporte es simplificado, solo generar el archivo básico
            try:
                if tipo_reporte == "Unificado":
                    # Para el reporte unificado
                    self.generar_excel_basico(configuraciones_sql, 
                                            consultas=consultas_bd, 
                                            nombres_consultas=nombres_consultas_bd, 
                                            consultas_mtj=consultas_mtj, 
                                            nombres_consultas_mtj=nombres_consultas_mtj, 
                                            nombre_archivo=nombre_archivo_basico, 
                                            unificado=True)
                elif tipo_reporte == "Observaciones BD":
                    self.generar_excel_basico(configuraciones_sql, 
                                            consultas=consultas, 
                                            nombres_consultas=nombres_consultas, 
                                            nombre_archivo=nombre_archivo_basico, 
                                            unificado=False)
                elif tipo_reporte == "BD MTJ Observaciones":
                    self.generar_excel_basico(configuraciones_sql, 
                                            consultas=consultas, 
                                            nombres_consultas=nombres_consultas, 
                                            nombre_archivo=nombre_archivo_basico, 
                                            unificado=False)

                else:
                    raise ValueError(f"Tipo de reporte no reconocido: {tipo_reporte}")
            except Exception as e:
                self.update_progress(f"Error en la generación del archivo básico: {e}")
                QMessageBox.critical(None, "Error", f"Error en la generación del archivo básico: {e}")
                return

           

        elif self.consolidado_radio.isChecked():
            # Si el reporte es consolidado, generar el archivo completo formateado
            writer = pd.ExcelWriter(nombre_archivo_final, engine='openpyxl')

            if tipo_reporte == "Observaciones BD":
                self.generar_hoja(writer, configuraciones_sql, consultas, nombres_consultas, "Observaciones BD")
            
            elif tipo_reporte == "BD MTJ Observaciones":
                self.generar_hoja(writer, configuraciones_sql, consultas, nombres_consultas, "BD MTJ Observaciones")

            elif tipo_reporte == "Unificado":
                # Generar hojas para el reporte unificado
                self.generar_hoja(writer, configuraciones_sql, consultas_bd, nombres_consultas_bd, "Observaciones BD")
                self.update_progress("Generando hoja BD MTJ Observaciones...")
                self.generar_hoja(writer, configuraciones_sql, consultas_mtj, nombres_consultas_mtj, "BD MTJ Observaciones")

            # Agregar la hoja "REV" como la primera
            #self.agregar_hoja_rev(writer, consultas, nombres_consultas)

            # Agregar la hoja "REV" según el tipo de reporte
            if tipo_reporte == "Unificado":
                self.agregar_hoja_rev_unificado(writer, consultas_bd, nombres_consultas_bd, consultas_mtj, nombres_consultas_mtj)

            elif tipo_reporte == "BD MTJ Observaciones" or tipo_reporte == "Observaciones BD":
                self.agregar_hoja_rev(writer, consultas, nombres_consultas)



            

            writer.close()

            # Verificar si el archivo se ha generado correctamente
            if os.path.exists(nombre_archivo_final):
                self.aplicar_filtros_excel(nombre_archivo_final)
                self.update_progress("Finalizado")

                # Crear un enlace HTML clicable
                mensaje_exito = f"""
                <html>
                <body>
                <p>Reportes generados exitosamente en <a href="file:///{nombre_archivo_final.replace(' ', '%20')}">{nombre_archivo_final}</a></p>
                </body>
                </html>
                """

                # Mostrar el mensaje de éxito con el enlace clicable
                msg = QMessageBox()
                msg.setWindowTitle("Éxito")
                msg.setText(mensaje_exito)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.setTextFormat(Qt.RichText)  # Permitir el formato HTML
                msg.exec_()
            else:
                QMessageBox.critical(None, "Error", "No se pudo generar el archivo de reporte.")



    def agregar_hoja_rev(self, writer, consultas, nombres_consultas):
        # Crear la hoja REV
        sheet_name = "REV"
        df_rev = pd.DataFrame(columns=["ID", "DESCRIPCIÓN REGLA", "CUMPLE", "NO CUMPLE", "OBSERVACIONES"])

        for consulta, nombre_consulta in zip(consultas, nombres_consultas):
            try:
                df_result = pd.read_sql_query(consulta, self.conexion)

                # Separar ID y DESCRIPCIÓN REGLA
                consulta_split = nombre_consulta.split("-", 1)
                id_consulta = consulta_split[0].strip() if len(consulta_split) > 1 else ""
                descripcion_regla = consulta_split[1].strip() if len(consulta_split) > 1 else nombre_consulta

                if df_result.empty:
                    df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "X", "NO CUMPLE": "", "OBSERVACIONES": ""}, ignore_index=True)
                else:
                    df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "", "NO CUMPLE": "X", "OBSERVACIONES": f"{len(df_result)} registros para verificar y/o ajustar"}, ignore_index=True)
            except Exception as e:
                self.update_progress(f"Error al ejecutar la consulta '{nombre_consulta}': {e}")
                df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "", "NO CUMPLE": "X", "OBSERVACIONES": "Error en consulta"}, ignore_index=True)

        # Escribir la hoja REV en el archivo Excel desde la fila 6
        df_rev.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)

        # Obtener la hoja recién creada
        worksheet = writer.sheets[sheet_name]

        # Formato de alineación: centrar el texto en las columnas específicas
        for cell in worksheet['A']:  # Columna ID
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['C']:  # Columna Cumple
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['D']:  # Columna No Cumple
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['B']:  # Columna DESCRIPCIÓN REGLA
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for cell in worksheet['E']:  # Columna Observaciones
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Obtener la ruta del plugin y agregar la imagen
        plugin_dir = os.path.dirname(__file__)  # Obtener la ruta donde se encuentra el plugin
        ruta_imagen = os.path.join(plugin_dir, 'ant.png')  # Ruta relativa a la imagen

        def cm_to_pixels(cm):
            return int(cm * 37.8)

        ancho_cm = 2.43  # Ancho en centímetros
        alto_cm = 2.43  # Alto en centímetros

        # Insertar la imagen en la columna A, filas 1-4
        img = Image(ruta_imagen)
        img.width = cm_to_pixels(ancho_cm)  # Ajusta el ancho de la imagen
        img.height = cm_to_pixels(alto_cm)  # Ajusta la altura de la imagen
        worksheet.merge_cells('A1:A4')  # Combinar las celdas de la columna A, filas 1-4
        worksheet.add_image(img, 'A1')  # Insertar la imagen en la celda A1

        # Colocar el texto en la columna B, filas 1 a 4
        worksheet.merge_cells('B1:C2')  # Combinar celdas para EQUIPO SIG - SPO
        worksheet['B1'] = "EQUIPO SIG - SPO"
        worksheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B1'].font = Font(bold=True)

        worksheet.merge_cells('B3:C3')  # Combinar celdas para RESULTADO VALIDACIÓN BASES DE DATOS
        worksheet['B3'] = "RESULTADO VALIDACIÓN BASES DE DATOS"
        worksheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B3'].font = Font(bold=True)

        worksheet.merge_cells('B4:C4')  # Combinar celdas para MUNICIPIO
        worksheet['B4'] = "MUNICIPIO:"
        worksheet['B4'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('D1:D2')  # Combinar celdas para Número de Revisión
        worksheet['D1'] = "Número de Revisión:"
        worksheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet['D3'] = "Fecha de Validación:"
        worksheet['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Insertar la fecha actual en la columna E, fila 3
        fecha_revision = datetime.datetime.now().strftime("%d/%m/%Y")
        worksheet['E3'] = fecha_revision
        worksheet['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells('E1:E2')  # Combinar celdas para número de revisión
        worksheet['E1'] = "1"
        worksheet['E1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet['E1'].font = Font(bold=True, size=13)
        worksheet['E1'].fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

        worksheet.merge_cells('D4:E4')  # Combinar celdas para validadores
        worksheet['D4'] = "VALIDADORES VERSION: CONSULTA_VALIDADORES_V4.3.sql"
        worksheet['D4'].alignment = Alignment(horizontal='center', vertical='center')

        # En la fila 5, combinar desde la columna A hasta la columna E
        worksheet.merge_cells('A5:E5')
        worksheet['A5'] = "1. REGLAS DE CONSISTENCIA LÓGICA SOBRE LA BD"
        worksheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A5'].font = Font(bold=True)
        worksheet['A5'].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

        # Ajustar los anchos de las columnas
        worksheet.column_dimensions['A'].width = 15  # Ancho para ID
        worksheet.column_dimensions['B'].width = 100  # Ancho para DESCRIPCIÓN REGLA
        worksheet.column_dimensions['C'].width = 15  # Cumple
        worksheet.column_dimensions['D'].width = 15  # No Cumple
        worksheet.column_dimensions['E'].width = 50  # Observaciones


        # Post-procesamiento para agregar títulos dinámicos
        titulos_grupo = {
            "2": "Cantidad de Registros",
            "3": "Verificación de duplicidad en identificadores únicos",
            "4": "Verificación de condición del predio y Tipo de derecho",
            "5": "Validación de fechas",
            "6": "Validación tipo de fuentes administrativas",
            "7": "Validación de relaciones lógicas",
            "8": "Validación de la coherencia en la informacion del interesado",
            "9": "Validación de diligenciento de las caracteristicas de unidad de construccion",
            "10": "Datos por diligenciar en el FILI",
            "11": "TOPOLOGÍA CONSTRUCCIONES",
            "12": "TOPOLOGÍA TERRENO",
            "13": "Areas Base de Datos - MTJ",
            "14": "Coindicencia Fuente administrativa Base de Datos - MTJ",
            "15": "Coincidencia de formal/informal entre Base de Datos y MTJ",
            "16": "la cantidad de posesiones y ocupaciones debe se igual en MTJ (COLUMNA BS) y BD",
            "17": "Coincidencia QR, FMI,  números prediales MTJ y Base de Datos",
            "18": "Coincidencia Interesado MTJ y Base de Datos",
            "19": "Completitud de llenado de campos en la MTJ"
            
        }

        rows_to_insert = []
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and ".1" in id_consulta:
                id_grupo = id_consulta.split(".")[0]
                if id_grupo in titulos_grupo:
                    rows_to_insert.append((row, id_grupo, titulos_grupo[id_grupo]))

        for row, id_grupo, titulo in reversed(rows_to_insert):
            worksheet.insert_rows(row)
            worksheet[f"A{row}"] = id_grupo
            worksheet[f"B{row}"] = titulo
            worksheet[f"A{row}"].font = Font(bold=True)
            worksheet[f"B{row}"].font = Font(bold=True)
            # Centrar el texto en las filas insertadas
            worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[f"B{row}"].alignment = Alignment(horizontal='center', vertical='center')





        # Agregar nuevas filas antes del ID 2
        rows_to_insert_before_2 = [
            ("1", "Verificación cargue XTF - .SQL"),
            ("1.1", "Cargue de información a base de datos con el XTF por medio del ILI SUITE"),
            ("1.2", "Generar proyecto en QGIS desde model baker")
        ]


        # Buscar la fila donde está el ID 2 y agregar las nuevas filas antes de esa
        # Buscar la fila donde está el ID 2 y agregar las nuevas filas antes de esa
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "2":
                for new_id, new_description in reversed(rows_to_insert_before_2):
                    worksheet.insert_rows(row)
                    worksheet[f"A{row}"] = new_id
                    worksheet[f"B{row}"] = new_description
                    
                    # El ID siempre centrado
                    worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Descripción centrada y en negrilla solo si es el ID 1
                    if new_id == "1":
                        worksheet[f"A{row}"].font = Font(bold=True)
                        worksheet[f"B{row}"].font = Font(bold=True)
                        worksheet[f"B{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar las nuevas filas





        # Lista de filas a insertar alrededor del ID 12.4
        rows_to_insert_around_12_4 = [
            ("12.3", "Predios formales adyacentes no pueden tener vacíos entre ellos"),  # Antes del 12.4
            ("12.5", "Predios informales adyacentes no se pueden tener vacíos entre ellos"),  # Después del 12.4
            ("12.6", "Las informalidades deben estar contenidas dentro de su formalidad y sin errores topológicos"),  # Después del 12.5
            ("12.7", "Las informalidades NO deben superar el límite de su formalidad")  # Después del 12.6
        ]

        # Buscar la fila donde está el ID 12.4 y agregar las nuevas filas antes y después
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "12.4":
                # Insertar "12.3" antes del 12.4
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "12.3"
                worksheet[f"B{row}"] = "Predios formales adyacentes no pueden tener vacíos entre ellos"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')

                # Ajustar la fila del 12.4 para evitar vacíos
                row += 1

                # Insertar las filas después del 12.4 (12.5, 12.6, 12.7)
                for new_id, new_description in rows_to_insert_around_12_4[1:]:
                    row += 1  # Incrementar la fila para cada inserción
                    worksheet.insert_rows(row)
                    worksheet[f"A{row}"] = new_id
                    worksheet[f"B{row}"] = new_description
                    worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar las filas necesarias







        # Buscar la fila donde está el ID 2.1 y agregar la nueva fila después de esa
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "2.2":
                # Insertar la nueva fila después del 2.1
                row += 1  # Mover a la siguiente fila para insertar
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "2.3"
                worksheet[f"B{row}"] = "La cantidad de soporte documentales debe ser igual a la cantidad de registros en cca_predio"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar la nueva fila


        # Agregar título y encabezado antes del ID 11
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "11":
                # Insertar título
                worksheet.insert_rows(row)
                worksheet.merge_cells(f"A{row}:E{row}")
                worksheet[f"A{row}"] = "2. REGLAS DE VALIDACIÓN TOPOLOGICA"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"A{row}"].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

                # Insertar encabezados
                row += 1
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "ID"
                worksheet[f"B{row}"] = "DESCRIPCIÓN REGLA"
                worksheet[f"C{row}"] = "CUMPLE"
                worksheet[f"D{row}"] = "NO CUMPLE"
                worksheet[f"E{row}"] = "OBSERVACIONES"

                for col in ["A", "B", "C", "D", "E"]:
                    worksheet[f"{col}{row}"].font = Font(bold=True)
                    worksheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

                break  # Detener después de insertar encabezados y título antes del ID 11





            

        # Agregar título y encabezado antes del ID 11
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "13":
                # Insertar título
                worksheet.insert_rows(row)
                worksheet.merge_cells(f"A{row}:E{row}")
                worksheet[f"A{row}"] = "3. CORRESPONDENCIA DE INFORMACIÓN ENTRE BASE DE DATOS Y  MTJ"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"A{row}"].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

                # Insertar encabezados
                row += 1
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "ID"
                worksheet[f"B{row}"] = "DESCRIPCIÓN REGLA"
                worksheet[f"C{row}"] = "CUMPLE"
                worksheet[f"D{row}"] = "NO CUMPLE"
                worksheet[f"E{row}"] = "OBSERVACIONES"

                for col in ["A", "B", "C", "D", "E"]:
                    worksheet[f"{col}{row}"].font = Font(bold=True)
                    worksheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

                break  # Detener después de insertar encabezados y título antes del ID 11



        # Definir el estilo de borde
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Aplicar el borde a todas las celdas con información
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border


        # Mover la hoja "REV" a la primera posición
        writer.book._sheets = [worksheet] + [s for s in writer.book._sheets if s != worksheet]

        # Guardar el archivo
        writer.save()






    def agregar_hoja_rev_unificado(self, writer, consultas_bd, nombres_consultas_bd, consultas_mtj, nombres_consultas_mtj):
        # Crear la hoja REV
        sheet_name = "REV"
        df_rev = pd.DataFrame(columns=["ID", "DESCRIPCIÓN REGLA", "CUMPLE", "NO CUMPLE", "OBSERVACIONES"])

        # Procesar todas las consultas, incluyendo BD y MTJ
        for consulta, nombre_consulta in zip(consultas_bd + consultas_mtj, nombres_consultas_bd + nombres_consultas_mtj):
            try:
                df_result = pd.read_sql_query(consulta, self.conexion)

                # Separar ID y DESCRIPCIÓN REGLA
                consulta_split = nombre_consulta.split("-", 1)
                id_consulta = consulta_split[0].strip() if len(consulta_split) > 1 else ""
                descripcion_regla = consulta_split[1].strip() if len(consulta_split) > 1 else nombre_consulta

                # Verificar si la consulta tiene resultados
                if df_result.empty:
                    df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "X", "NO CUMPLE": "", "OBSERVACIONES": ""}, ignore_index=True)
                else:
                    df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "", "NO CUMPLE": "X", "OBSERVACIONES": f"{len(df_result)} registros para verificar y/o ajustar"}, ignore_index=True)
            except Exception as e:
                self.update_progress(f"Error al ejecutar la consulta '{nombre_consulta}': {e}")
                df_rev = df_rev.append({"ID": id_consulta, "DESCRIPCIÓN REGLA": descripcion_regla, "CUMPLE": "", "NO CUMPLE": "X", "OBSERVACIONES": "Error en consulta"}, ignore_index=True)

        # Escribir la hoja REV en el archivo Excel desde la fila 6
        df_rev.to_excel(writer, sheet_name=sheet_name, index=False, startrow=5)

        # Obtener la hoja recién creada
        worksheet = writer.sheets[sheet_name]

        # En la fila 5, combinar desde la columna A hasta la columna E
        worksheet.merge_cells('A5:E5')
        worksheet['A5'] = "1. REGLAS DE CONSISTENCIA LÓGICA SOBRE LA BD"
        worksheet['A5'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A5'].font = Font(bold=True)
        worksheet['A5'].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

        # Ajustar los anchos de las columnas
        worksheet.column_dimensions['A'].width = 15  # Ancho para ID
        worksheet.column_dimensions['B'].width = 100  # Ancho para DESCRIPCIÓN REGLA
        worksheet.column_dimensions['C'].width = 15  # Cumple
        worksheet.column_dimensions['D'].width = 15  # No Cumple
        worksheet.column_dimensions['E'].width = 50  # Observaciones




        # Alinear el texto en las columnas específicas
        for cell in worksheet['A']:  # Columna ID
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['C']:  # Columna Cumple
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['D']:  # Columna No Cumple
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet['B']:  # Columna DESCRIPCIÓN REGLA
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        for cell in worksheet['E']:  # Columna Observaciones
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Agregar la imagen del encabezado
        plugin_dir = os.path.dirname(__file__)  # Obtener la ruta donde se encuentra el plugin
        ruta_imagen = os.path.join(plugin_dir, 'ant.png')  # Ruta relativa a la imagen

        def cm_to_pixels(cm):
            return int(cm * 37.8)

        ancho_cm = 2.43  # Ancho en centímetros
        alto_cm = 2.43  # Alto en centímetros

        # Insertar la imagen en la columna A, filas 1-4
        img = Image(ruta_imagen)
        img.width = cm_to_pixels(ancho_cm)  # Ajusta el ancho de la imagen
        img.height = cm_to_pixels(alto_cm)  # Ajusta la altura de la imagen
        worksheet.merge_cells('A1:A4')  # Combinar las celdas de la columna A, filas 1-4
        worksheet.add_image(img, 'A1')  # Insertar la imagen en la celda A1

        # Colocar el texto en la columna B, filas 1 a 4
        worksheet.merge_cells('B1:C2')  # Combinar celdas para EQUIPO SIG - SPO
        worksheet['B1'] = "EQUIPO SIG - SPO"
        worksheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B1'].font = Font(bold=True)

        worksheet.merge_cells('B3:C3')  # Combinar celdas para RESULTADO VALIDACIÓN BASES DE DATOS
        worksheet['B3'] = "RESULTADO VALIDACIÓN BASES DE DATOS"
        worksheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['B3'].font = Font(bold=True)

        worksheet.merge_cells('B4:C4')  # Combinar celdas para MUNICIPIO
        worksheet['B4'] = "MUNICIPIO:"
        worksheet['B4'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet.merge_cells('D1:D2')  # Combinar celdas para Número de Revisión
        worksheet['D1'] = "Número de Revisión:"
        worksheet['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet['D3'] = "Fecha de Validación:"
        worksheet['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Insertar la fecha actual en la columna E, fila 3
        fecha_revision = datetime.datetime.now().strftime("%d/%m/%Y")
        worksheet['E3'] = fecha_revision
        worksheet['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells('E1:E2')  # Combinar celdas para número de revisión
        worksheet['E1'] = "1"
        worksheet['E1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet['E1'].font = Font(bold=True, size=13)
        worksheet['E1'].fill = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

        worksheet.merge_cells('D4:E4')  # Combinar celdas para validadores
        worksheet['D4'] = "VALIDADORES VERSION: CONSULTA_VALIDADORES_V4.3.sql"
        worksheet['D4'].alignment = Alignment(horizontal='center', vertical='center')



        titulos_grupo = {
            "2": "Cantidad de Registros",
            "3": "Verificación de duplicidad en identificadores únicos",
            "4": "Verificación de condición del predio y Tipo de derecho",
            "5": "Validación de fechas",
            "6": "Validación tipo de fuentes administrativas",
            "7": "Validación de relaciones lógicas",
            "8": "Validación de la coherencia en la informacion del interesado",
            "9": "Validación de diligenciento de las caracteristicas de unidad de construccion",
            "10": "Datos por diligenciar en el FILI",
            "11": "TOPOLOGÍA CONSTRUCCIONES",
            "12": "TOPOLOGÍA TERRENO",
            "13": "Areas Base de Datos - MTJ",
            "14": "Coindicencia Fuente administrativa Base de Datos - MTJ",
            "15": "Coincidencia de formal/informal entre Base de Datos y MTJ",
            "16": "la cantidad de posesiones y ocupaciones debe se igual en MTJ (COLUMNA BS) y BD",
            "17": "Coincidencia QR, FMI,  números prediales MTJ y Base de Datos",
            "18": "Coincidencia Interesado MTJ y Base de Datos",
            "19": "Completitud de llenado de campos en la MTJ"
            
        }

        rows_to_insert = []
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and ".1" in id_consulta:
                id_grupo = id_consulta.split(".")[0]
                if id_grupo in titulos_grupo:
                    rows_to_insert.append((row, id_grupo, titulos_grupo[id_grupo]))

        for row, id_grupo, titulo in reversed(rows_to_insert):
            worksheet.insert_rows(row)
            worksheet[f"A{row}"] = id_grupo
            worksheet[f"B{row}"] = titulo
            worksheet[f"A{row}"].font = Font(bold=True)
            worksheet[f"B{row}"].font = Font(bold=True)
            # Centrar el texto en las filas insertadas
            worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[f"B{row}"].alignment = Alignment(horizontal='center', vertical='center')






        # Agregar nuevas filas antes del ID 2
        rows_to_insert_before_2 = [
            ("1", "Verificación cargue XTF - .SQL"),
            ("1.1", "Cargue de información a base de datos con el XTF por medio del ILI SUITE"),
            ("1.2", "Generar proyecto en QGIS desde model baker")
        ]


        # Buscar la fila donde está el ID 2 y agregar las nuevas filas antes de esa
        # Buscar la fila donde está el ID 2 y agregar las nuevas filas antes de esa
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "2":
                for new_id, new_description in reversed(rows_to_insert_before_2):
                    worksheet.insert_rows(row)
                    worksheet[f"A{row}"] = new_id
                    worksheet[f"B{row}"] = new_description
                    
                    # El ID siempre centrado
                    worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Descripción centrada y en negrilla solo si es el ID 1
                    if new_id == "1":
                        worksheet[f"A{row}"].font = Font(bold=True)
                        worksheet[f"B{row}"].font = Font(bold=True)
                        worksheet[f"B{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar las nuevas filas






        # Lista de filas a insertar alrededor del ID 12.4
        rows_to_insert_around_12_4 = [
            ("12.3", "Predios formales adyacentes no pueden tener vacíos entre ellos"),  # Antes del 12.4
            ("12.5", "Predios informales adyacentes no se pueden tener vacíos entre ellos"),  # Después del 12.4
            ("12.6", "Las informalidades deben estar contenidas dentro de su formalidad y sin errores topológicos"),  # Después del 12.5
            ("12.7", "Las informalidades NO deben superar el límite de su formalidad")  # Después del 12.6
        ]

        # Buscar la fila donde está el ID 12.4 y agregar las nuevas filas antes y después
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "12.4":
                # Insertar "12.3" antes del 12.4
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "12.3"
                worksheet[f"B{row}"] = "Predios formales adyacentes no pueden tener vacíos entre ellos"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')

                # Ajustar la fila del 12.4 para evitar vacíos
                row += 1

                # Insertar las filas después del 12.4 (12.5, 12.6, 12.7)
                for new_id, new_description in rows_to_insert_around_12_4[1:]:
                    row += 1  # Incrementar la fila para cada inserción
                    worksheet.insert_rows(row)
                    worksheet[f"A{row}"] = new_id
                    worksheet[f"B{row}"] = new_description
                    worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                    worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar las filas necesarias

        # Buscar la fila donde está el ID 2.1 y agregar la nueva fila después de esa
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "2.2":
                # Insertar la nueva fila después del 2.1
                row += 1  # Mover a la siguiente fila para insertar
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "2.3"
                worksheet[f"B{row}"] = "La cantidad de soporte documentales debe ser igual a la cantidad de registros en cca_predio"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"B{row}"].alignment = Alignment(horizontal='left', vertical='center')
                break  # Detener después de insertar la nueva fila



        # Agregar título y encabezado antes del ID 11
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "11":
                # Insertar título
                worksheet.insert_rows(row)
                worksheet.merge_cells(f"A{row}:E{row}")
                worksheet[f"A{row}"] = "2. REGLAS DE VALIDACIÓN TOPOLOGICA"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"A{row}"].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

                # Insertar encabezados
                row += 1
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "ID"
                worksheet[f"B{row}"] = "DESCRIPCIÓN REGLA"
                worksheet[f"C{row}"] = "CUMPLE"
                worksheet[f"D{row}"] = "NO CUMPLE"
                worksheet[f"E{row}"] = "OBSERVACIONES"

                for col in ["A", "B", "C", "D", "E"]:
                    worksheet[f"{col}{row}"].font = Font(bold=True)
                    worksheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

                break  # Detener después de insertar encabezados y título antes del ID 11




        # Agregar título y encabezado antes del ID 11
        for row in range(6, worksheet.max_row + 1):
            id_consulta = worksheet[f"A{row}"].value
            if id_consulta and id_consulta == "13":
                # Insertar título
                worksheet.insert_rows(row)
                worksheet.merge_cells(f"A{row}:E{row}")
                worksheet[f"A{row}"] = "3. CORRESPONDENCIA DE INFORMACIÓN ENTRE BASE DE DATOS Y  MTJ"
                worksheet[f"A{row}"].alignment = Alignment(horizontal='center', vertical='center')
                worksheet[f"A{row}"].font = Font(bold=True)
                worksheet[f"A{row}"].fill = PatternFill(start_color="DAF2D0", end_color="DAF2D0", fill_type="solid")

                # Insertar encabezados
                row += 1
                worksheet.insert_rows(row)
                worksheet[f"A{row}"] = "ID"
                worksheet[f"B{row}"] = "DESCRIPCIÓN REGLA"
                worksheet[f"C{row}"] = "CUMPLE"
                worksheet[f"D{row}"] = "NO CUMPLE"
                worksheet[f"E{row}"] = "OBSERVACIONES"

                for col in ["A", "B", "C", "D", "E"]:
                    worksheet[f"{col}{row}"].font = Font(bold=True)
                    worksheet[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

                break  # Detener después de insertar encabezados y título antes del ID 11






        # Definir el estilo de borde
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Aplicar el borde a todas las celdas con información
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border


        # Mover la hoja "REV" a la primera posición
        writer.book._sheets = [worksheet] + [s for s in writer.book._sheets if s != worksheet]

        # Guardar el archivo
        writer.save()








    def generar_excel_basico(self, configuraciones_sql, consultas=None, nombres_consultas=None, consultas_mtj=None, nombres_consultas_mtj=None, nombre_archivo=None, unificado=False):
        cursor = self.conexion.cursor()
        for configuracion in configuraciones_sql:
            cursor.execute(configuracion)
        self.conexion.commit()

        try:
            # Iniciar la creación del archivo Excel
            writer = pd.ExcelWriter(nombre_archivo, engine='xlsxwriter')
            workbook = writer.book
            self.update_progress(f"Generando archivo Excel básico: {nombre_archivo}")

            # Si es unificado, procesar los dos conjuntos de consultas
            if unificado:
                self.update_progress(f"Generando hojas para el reporte unificado")
                
                # Generar hojas para consultas del reporte 1 (color verde brillante #01FF13)
                for i, (consulta, nombre_consulta) in enumerate(zip(consultas, nombres_consultas), start=1):
                    if consulta and nombre_consulta:
                        self.generar_hoja_excel_basico(writer, consulta, nombre_consulta, "#01FF13")
                
                # Generar hojas para consultas del reporte 2 (color amarillo claro #ECF10F)
                for i, (consulta, nombre_consulta) in enumerate(zip(consultas_mtj, nombres_consultas_mtj), start=1):
                    if consulta and nombre_consulta:
                        self.generar_hoja_excel_basico(writer, consulta, nombre_consulta, "#ECF10F")

            else:
                self.update_progress(f"Generando hojas para reportes individuales")
                # Generar hojas para reportes individuales (sin colores)
                for i, (consulta, nombre_consulta) in enumerate(zip(consultas, nombres_consultas), start=1):
                    if consulta and nombre_consulta:
                        self.generar_hoja_excel_basico(writer, consulta, nombre_consulta)

            # Guardar el archivo
            self.update_progress(f"Guardando archivo Excel básico: {nombre_archivo}")
            writer.close()

            # Mostrar mensaje de éxito al finalizar la generación del Excel básico
            mensaje_exito = f"""
            <html>
            <body>
            <p>El archivo Excel básico ha sido generado exitosamente en <a href="file:///{nombre_archivo.replace(' ', '%20')}">{nombre_archivo}</a></p>
            </body>
            </html>
            """
            msg = QMessageBox()
            msg.setWindowTitle("Éxito - Excel Básico")
            msg.setText(mensaje_exito)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.setTextFormat(Qt.RichText)  # Permitir el formato HTML
            msg.exec_()

        except Exception as e:
            self.update_progress(f"Error al generar el archivo básico: {e}")
            QMessageBox.critical(None, "Error", f"Error al generar el archivo básico: {e}")
            return






    def generar_hoja_excel_basico(self, writer, consulta, nombre_consulta, color_pestaña=None):
        try:
            self.update_progress(f"Procesando consulta '{nombre_consulta}'...")

            # Truncar el nombre de la hoja a 31 caracteres
            nombre_consulta_truncado = nombre_consulta[:31]

            # Ejecutar la consulta y obtener el DataFrame
            self.update_progress(f"Ejecutando consulta para '{nombre_consulta}'...")
            df = pd.read_sql_query(consulta, self.conexion)

            if df.empty:
                self.update_progress(f"La consulta '{nombre_consulta}' no generó resultados.")
                return

            # Crear hoja en el Excel
            self.update_progress(f"Creando hoja para '{nombre_consulta_truncado}'")
            df.to_excel(writer, sheet_name=nombre_consulta_truncado, index=False, startrow=1)

            # Verificar si la hoja se creó correctamente
            worksheet = writer.sheets.get(nombre_consulta_truncado)
            if worksheet is None:
                raise Exception(f"No se pudo crear la hoja: {nombre_consulta_truncado}")
            else:
                self.update_progress(f"Hoja creada correctamente para '{nombre_consulta_truncado}'")

            # Escribir el nombre completo de la consulta en la primera celda
            worksheet.write(0, 0, nombre_consulta)

            # Aplicar color si se proporciona
            if color_pestaña:
                worksheet.set_tab_color(color_pestaña)  # Aplicar el color personalizado a la pestaña

            # Autoajustar el ancho de las columnas
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.set_column(idx, idx, max_len)

        except Exception as e:
            self.update_progress(f"Error al procesar la consulta '{nombre_consulta}': {e}")
            QMessageBox.critical(None, "Error", f"Error al procesar la consulta '{nombre_consulta}': {e}")
            return

    






    def generar_hoja(self, writer, configuraciones_sql, consultas, nombres_consultas, hoja_nombre):
        cursor = self.conexion.cursor()
        for configuracion in configuraciones_sql:
            cursor.execute(configuracion)
        self.conexion.commit()

        columna_inicio = 0
        total_consultas = len(consultas)

        self.update_progress(f"Generando hoja {hoja_nombre}...")
        for i, (consulta, nombre_consulta) in enumerate(zip(consultas, nombres_consultas), start=1):
            try:
                self.update_progress(f"Ejecutando consulta {i}: {nombre_consulta}...")
                subconsultas = consulta.strip().split(";")
                df_concat = None

                for subconsulta in subconsultas:
                    if subconsulta.strip():
                        df_sub = pd.read_sql_query(subconsulta, self.conexion)
                        if not df_sub.empty:
                            if df_concat is None:
                                df_concat = df_sub
                            else:
                                df_concat = pd.concat([df_concat, pd.DataFrame([[''] * df_sub.shape[1]], columns=df_sub.columns), df_sub], ignore_index=True)

                if df_concat is not None:
                    df_concat.columns = map(str.upper, df_concat.columns)
                    df_concat.to_excel(writer, sheet_name=hoja_nombre, startcol=columna_inicio, startrow=2, index=False, header=True)
                    columna_inicio += df_concat.shape[1] + 2




            except Exception as e:
                self.update_progress(f"Error al ejecutar la consulta '{nombre_consulta}': {e}")
                QMessageBox.critical(None, "Error", f"Error al ejecutar la consulta '{nombre_consulta}': {e}")
                return

        self.update_progress(f"Aplicando formato a la hoja {hoja_nombre}...")
        self.format_hoja(writer, hoja_nombre, consultas, nombres_consultas)
        self.update_progress(f"Finalizando hoja {hoja_nombre}...")





    def format_hoja(self, writer, hoja_nombre, consultas, nombres_consultas):
        book = writer.book
        sheet = book[hoja_nombre]
        columna_inicio = 0
        total_consultas = len(consultas)

        for i, (consulta, nombre_consulta) in enumerate(zip(consultas, nombres_consultas), start=1):
            try:
                self.update_progress(f"Aplicando formato a la consulta {i}: {nombre_consulta} en la hoja {hoja_nombre}...")
                subconsultas = consulta.strip().split(";")
                df_concat = None

                for subconsulta in subconsultas:
                    if subconsulta.strip():
                        df_sub = pd.read_sql_query(subconsulta, self.conexion)
                        if not df_sub.empty:
                            if df_concat is None:
                                df_concat = df_sub
                            else:
                                df_concat = pd.concat([df_concat, pd.DataFrame([[''] * df_sub.shape[1]], columns=df_sub.columns), df_sub], ignore_index=True)

                if df_concat is not None:
                    col_inicio = columna_inicio + 1
                    col_fin = columna_inicio + df_concat.shape[1]

                    sheet.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_fin)
                    cell = sheet.cell(row=1, column=col_inicio)
                    cell.value = nombre_consulta
                    cell.alignment = Alignment(horizontal='center',wrap_text=True)
                    sheet.row_dimensions[1].height = 50

                    # Asegurar que todas las celdas de la fila 1 tengan ajuste de texto activado
                    for col in range(col_inicio, col_fin + 1):
                        cell = sheet.cell(row=1, column=col)
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)

                    # Ajustar la altura de la fila 3 a 50 píxeles
                    sheet.row_dimensions[3].height = 50

                    # Asegurar que todas las celdas de la fila 3 tengan ajuste de texto activado
                    for col in range(col_inicio, col_fin + 1):
                        cell = sheet.cell(row=3, column=col)
                        cell.alignment = Alignment(wrap_text=True)


                    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=col_inicio, max_col=col_fin):
                        for cell in row:
                            cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

                    for row in sheet.iter_rows(min_row=2, min_col=col_inicio, max_row=sheet.max_row, max_col=col_fin):
                        for cell in row:
                            cell.border = Border(left=Side(style='thin', color='000000'),
                                                right=Side(style='thin', color='000000'),
                                                top=Side(style='thin', color='000000'),
                                                bottom=Side(style='thin', color='000000'))

                    for column_cells in sheet.columns:
                        column = [cell for cell in column_cells]
                        length = max(len(str(cell.value)) for cell in column)
                        col_letter = get_column_letter(column[0].column)
                        if length > 25:
                            sheet.column_dimensions[col_letter].width = 25
                            for cell in column:
                                cell.alignment = Alignment(wrap_text=True)
                        else:
                            sheet.column_dimensions[col_letter].width = length + 5

                    for row in sheet.iter_rows():
                        for cell in row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                    columna_inicio += df_concat.shape[1] + 2




            except Exception as e:
                self.update_progress(f"Error al procesar la consulta '{nombre_consulta}': {e}")
                QMessageBox.critical("Error", f"Error al procesar la consulta '{nombre_consulta}': {e}")
                return




 
    def aplicar_filtros_excel(self, nombre_archivo):
        self.update_progress("Aplicando filtros en el archivo Excel...")
        # Cargar el archivo Excel y aplicar filtros a cada hoja sobre la fila 3
        wb = load_workbook(nombre_archivo)
        for sheet_name in wb.sheetnames:
            if sheet_name == "REV":
                continue  # Saltar la hoja REV

            sheet = wb[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column

            # Aplicar el filtro desde la fila 3 hasta la última fila
            filter_range = f"A3:{get_column_letter(max_col)}{max_row}"
            sheet.auto_filter.ref = filter_range

            # Congelar la fila 3
            sheet.freeze_panes = sheet["A4"]

            # (Opcional) Proteger la hoja pero permitir el uso de filtros y sort
            # sheet.protection.sheet = True
            # sheet.protection.autoFilter = True
            # sheet.protection.sort = True
            # sheet.protection.insertRows = True
            # sheet.protection.insertColumns = True
            # sheet.protection.formatCells = True
            # Agregar una contraseña para proteger la hoja --sheet.protection.set_password('your_password_here')

        # Guardar el archivo Excel con los filtros aplicados
        wb.save(nombre_archivo)
        self.update_progress("Filtros aplicados y archivo guardado.")



    def update_progress(self, status):
        # Agregar el texto de progreso al QTextEdit
        self.status_text.append(status)

        # Verificar si el scroll ya está en la parte inferior
        scrollbar = self.status_text.verticalScrollBar()
        at_bottom = scrollbar.value() == scrollbar.maximum()

        # Si está en la parte inferior, hacemos scroll automático
        if at_bottom:
            scrollbar.setValue(scrollbar.maximum())

        # Permitir que la GUI se actualice mientras
        QApplication.processEvents()



def mostrar_error(error):
    app = QApplication([])  # Crea una instancia de QApplication
    ventana_error = QMainWindow()  # Crea la ventana principal
    ventana_error.setWindowTitle("Error")
    ventana_error.setGeometry(100, 100, 400, 200)

    # Crear un widget central y una área de desplazamiento
    widget_central = QWidget()
    layout = QVBoxLayout(widget_central)

    # Crear el QTextEdit para mostrar el error con scroll
    text_area = QTextEdit()
    text_area.setText(error)
    text_area.setReadOnly(True)  # Deshabilitar edición

    # Agregar el QTextEdit al layout
    layout.addWidget(text_area)

    # Establecer el widget central en la ventana
    ventana_error.setCentralWidget(widget_central)

    ventana_error.show()  # Mostrar la ventana
    app.exec_()  # Ejecutar la aplicación

if __name__ == "__main__":
    mostrar_error("Este es un mensaje de error ejemplo.")
