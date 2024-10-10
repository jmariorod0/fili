import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from PyQt5.QtWidgets import QFileDialog, QMessageBox
import csv
import datetime

def ejecutar_estandarizacion(update_progress):
    # Pedir al usuario que seleccione el archivo a convertir
    input_file_path, _ = QFileDialog.getOpenFileName(None, "Seleccione el archivo a convertir", "", "Excel files (*.xlsx)")
    if not input_file_path:
        QMessageBox.warning(None, "Error", "No se seleccionó ningún archivo.")
        return

    update_progress(f"Archivo seleccionado: {input_file_path}")

    # Leer el archivo Excel y las hojas "MTJ" y "1_INTERESADOS"
    update_progress("Leyendo la hoja 'MTJ' del archivo Excel...")
    df_mtj = pd.read_excel(input_file_path, sheet_name='MTJ', header=4)
    update_progress("Leyendo la hoja '1_INTERESADOS' del archivo Excel...")
    df_interesados = pd.read_excel(input_file_path, sheet_name='1_INTERESADOS', header=0).iloc[:, :18]

    # Verificar el número de columnas y seleccionar las primeras 124 si hay más
    max_columns_mtj = 124
    if df_mtj.shape[1] > max_columns_mtj:
        extra_columns = df_mtj.shape[1] - max_columns_mtj
        extra_column_names = df_mtj.columns[max_columns_mtj:]
        df_mtj = df_mtj.iloc[:, :max_columns_mtj]
        update_progress(f"Advertencia: El archivo de entrada tiene más de {max_columns_mtj} columnas. Se consideraron solo las primeras {max_columns_mtj}. No se consideraron las columnas adicionales: {', '.join(extra_column_names)}")

    # Nuevos nombres de encabezados para la hoja "MTJ"
    new_headers_mtj = [
        "ID", "uit", "local_id", "id_operacion", "fecha_asig", "profesional_tecnico", "fecha_analisi_tecnico", "estado_tecnico", 
        "observaciones_tecnico", "CR_InteresadoTipo", "CR_DocumentoTipo", "CR_Interesado_all", "Documento_Identidad", "jefe_hogar", 
        "victima_conflicto", "estado_civil", "inicio_tenencia", "reside_predio", "persona_distinta_reside", "quien", "explota_predio", 
        "metodo_levan", "area_terreno_levantamiento", "nombre_predio", "destino_economico", "tiene_rezago", "num_rezago", 
        "area_lpp_resgitral", "area_dif_catas_resgi", "tiene_construcciones", "cons_cantidad", "cons_area", "departamento", 
        "municipio", "vereda", "numero_predial", "numero_predial_provisional", "propietario_catstral", "direccion_catastral", 
        "area_catastral", "area_catastral_poligono", "area_cons_catastral", "destino_economico_catastral", "fecha_consulta_catastral", 
        "tipo_novedad", "complemneto_tipo_novedad", "clasificacion_suelo", "uso_suelo_plan_parcial", "categ_suelo_rural", 
        "categoria_proteccion", "desarrollo_restringido", "fecha_certificacion_uso_suelo", "cruza_determinantes", "cruza_desc_restricciones", 
        "cruza_desc_restricciones_area_por", "cruza_desc_condicionantes", "cruza_desc_condicionantes_area_por", "cruza_area_habilitada", 
        "concepto_catastral_general", "concepto_catastral_especifico", "req_acta_colindancia", "req_uso_suelo", "res_certificado_riesgos", 
        "profesional_juridico", "fecha_analisi_juridico", "estado_juridico", "observaciones_juridico", "formal_informal", "naturaleza_predio", 
        "tipo_predio", "tipo_derecho", "condicion_predio", "relaciona_fmi", "Codigo_Orip", "Matricula_Inmobiliaria", 
        "ExtReferenciaRegistralSistemaAntiguo", "fmi_propietario", "fmi_ref_catastral", "fmi_area", "fmi_direccion", "fmi_tipo_predio", 
        "fmi_matriz", "fmi_derivados", "tipo_fuente_adm", "fmi_ente_emisor", "numero_fuente_adm", "fmi_fecha_documento_fuente", 
        "disponibilidad_fuente_adm", "documento_apertura_fmi", "fecha_doc_acto_apertura", "fmi_medidas_cautelares", "fmi_salvedades_correc", 
        "fmi_complementa", "fmi_cabida_lindero", "fmi_titulo_originario", "fmi_estado", "tipo_novedad_fmi", "estado_rtdaf", "rupta", 
        "predio_cruza_pretension", "tipo_pretension", "medida_2333", "fadc", "req_ospr", "concepto_juridico", "tipologia_titulacion", 
        "ruta_atencion", "fmi_procedimientos_catatrales", "material_cimple", "doc_faltantes", "habilitado_reso", "reso", "profesional_agro", 
        "fecha_analisi_agro", "estado_agro", "observaciones_agro", "agro_metodo_analisis", "agro_ufh", "agro_cober_agro", "agrp_fuente_info", 
        "agro_aptitud", "agro_cultivos_ilicitos", "agro_recomendaciones", "agro_concepto"
    ]

    # Nuevos nombres de encabezados para la hoja "1_INTERESADOS"
    new_headers_interesados = [
        "qr_derivado", "interesado_tipo", "documento_tipo", "primer_nombre", "segundo_nombre", "primer_apellido", "segundo_apellido", 
        "interesado_concat", "documento_numero", "jefe_hogar", "victima_conflicto", "estado_civil", "fecha_inicio_tenencia", 
        "reside_predio_interesado", "reside_distinta_interesado", "quien", "explota_directamente", "numero_formulario_reso"
    ]

    # Renombrar los encabezados de los DataFrames
    update_progress("Renombrando encabezados de la hoja 'MTJ'...")
    df_mtj.columns = new_headers_mtj
    update_progress("Renombrando encabezados de la hoja '1_INTERESADOS'...")
    df_interesados.columns = new_headers_interesados

    # Reemplazar ';' con ',' en las columnas específicas
    update_progress("Reemplazando ';' con ',' en columnas específicas de 'MTJ'...")
    columns_to_replace = ['CR_InteresadoTipo', 'CR_DocumentoTipo', 'CR_Interesado_all', 'Documento_Identidad']
    for col in columns_to_replace:
        if col in df_mtj.columns:
            df_mtj[col] = df_mtj[col].str.replace(';', ',')

    # Filtrar filas donde 'local_id' y 'id_operacion' no están vacíos
    update_progress("Filtrando filas en 'MTJ' donde 'local_id' y 'id_operacion' no están vacíos...")
    df_mtj = df_mtj.dropna(subset=['local_id', 'id_operacion'])

    # Filtrar filas donde 'qr_derivado', 'interesado_tipo' y 'documento_tipo' no están vacíos en INTERESADOS
    update_progress("Filtrando filas en '1_INTERESADOS' donde 'qr_derivado', 'interesado_tipo' no están vacíos...")
    df_interesados = df_interesados.dropna(subset=['qr_derivado', 'interesado_tipo'])

    # Rellenar valores vacíos en la columna 'ID' con el mayor valor + 1
    update_progress("Rellenando valores vacíos en la columna 'ID'...")
    df_mtj['ID'] = pd.to_numeric(df_mtj['ID'], errors='coerce')
    max_id = df_mtj['ID'].max()
    df_mtj['ID'] = df_mtj['ID'].fillna(max_id + 1).astype(int)

    # Pedir al usuario que seleccione la carpeta de salida
    output_folder = QFileDialog.getExistingDirectory(None, "Seleccione la carpeta de salida")
    if not output_folder:
        QMessageBox.warning(None, "Error", "No se seleccionó ninguna carpeta de salida.")
        return

    update_progress("Carpeta de salida seleccionada: " + output_folder)

    # Obtener el nombre base del archivo de entrada sin la extensión
    base_filename = os.path.splitext(os.path.basename(input_file_path))[0]

    # Definir las rutas de salida para los archivos Excel y CSV
    output_file_path_mtj = os.path.join(output_folder, f"{base_filename}_MTJ_ESTANDARIZADA.xlsx")
    output_file_path_interesados = os.path.join(output_folder, f"{base_filename}_INTERESADOS_ESTANDARIZADO.xlsx")

    # Guardar los datos en los nuevos archivos Excel
    update_progress("Guardando los datos estandarizados en archivos Excel...")
    df_mtj.to_excel(output_file_path_mtj, index=False)
    df_interesados.to_excel(output_file_path_interesados, index=False)

    # Cargar el archivo Excel guardado para MTJ
    update_progress("Aplicando formato de fecha en la hoja 'MTJ'...")
    workbook_mtj = load_workbook(output_file_path_mtj)
    worksheet_mtj = workbook_mtj.active

    # Definir un estilo de fecha corta
    date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

    # Registrar el estilo en el workbook si no existe
    if 'date_style' not in workbook_mtj.named_styles:
        workbook_mtj.add_named_style(date_style)

    # Aplicar el estilo de fecha corta a las columnas específicas en MTJ
    date_columns_mtj = ['fecha_asig', 'fecha_analisi_tecnico', 'fecha_analisi_juridico','fmi_fecha_documento_fuente','fecha_analisi_agro','fecha_doc_acto_apertura','fecha_consulta_catastral']
    for col in date_columns_mtj:
        if col in df_mtj.columns:
            col_idx = df_mtj.columns.get_loc(col) + 1  # Obtener el índice de la columna (1-indexed)
            for row in range(2, len(df_mtj) + 2):  # Comenzar desde la fila 2 hasta el final de los datos
                cell = worksheet_mtj.cell(row=row, column=col_idx)
                if cell.value is not None:
                    try:
                        # Convertir la celda a fecha si es posible
                        cell.value = pd.to_datetime(cell.value).date()
                    except ValueError:
                        continue  # Ignorar si no es posible convertir a datetime
                    cell.number_format = 'DD/MM/YYYY'
                    cell.style = date_style

    # Guardar el archivo con las fechas formateadas para MTJ
    workbook_mtj.save(output_file_path_mtj)
    update_progress(f"Archivo Excel MTJ guardado exitosamente en: {output_file_path_mtj}")

    # Cargar el archivo Excel guardado para INTERESADOS
    update_progress("Aplicando formato de fecha en la hoja '1_INTERESADOS'...")
    workbook_interesados = load_workbook(output_file_path_interesados)
    worksheet_interesados = workbook_interesados.active

    # Aplicar el estilo de fecha corta a la columna específica en INTERESADOS
    date_columns_interesados = ['fecha_inicio_tenencia']
    for col in date_columns_interesados:
        if col in df_interesados.columns:
            col_idx = df_interesados.columns.get_loc(col) + 1  # Obtener el índice de la columna (1-indexed)
            for row in range(2, len(df_interesados) + 2):  # Comenzar desde la fila 2 hasta el final de los datos
                cell = worksheet_interesados.cell(row=row, column=col_idx)
                if cell.value is not None:
                    try:
                        # Convertir la celda a fecha si es posible
                        cell.value = pd.to_datetime(cell.value).date()
                    except ValueError:
                        continue  # Ignorar si no es posible convertir a datetime
                    cell.number_format = 'DD/MM/YYYY'
                    cell.style = date_style

    # Guardar el archivo con las fechas formateadas para INTERESADOS
    workbook_interesados.save(output_file_path_interesados)
    update_progress(f"Archivo Excel INTERESADOS guardado exitosamente en: {output_file_path_interesados}")

    # Definir las rutas de salida para los archivos CSV
    output_csv_path_mtj = os.path.join(output_folder, f"{base_filename}_MTJ_ESTANDARIZADA.csv")
    output_csv_path_interesados = os.path.join(output_folder, f"{base_filename}_INTERESADOS_ESTANDARIZADO.csv")

    # Función para guardar un archivo de Excel como CSV
    def save_excel_as_csv(excel_path, csv_path):
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        with open(csv_path, mode='w', newline='', encoding='utf-8-sig') as f:
            csv_writer = csv.writer(f, delimiter=';')
            for row in sheet.iter_rows(values_only=True):
                row = [cell.strftime('%d/%m/%Y') if isinstance(cell, (datetime.date, datetime.datetime)) else cell for cell in row]
                csv_writer.writerow(row)

    # Guardar los archivos Excel como CSV
    update_progress("Guardando los archivos Excel como CSV...")
    save_excel_as_csv(output_file_path_mtj, output_csv_path_mtj)
    save_excel_as_csv(output_file_path_interesados, output_csv_path_interesados)

    update_progress(f"Archivo CSV MTJ guardado exitosamente en: {output_csv_path_mtj}")
    update_progress(f"Archivo CSV INTERESADOS guardado exitosamente en: {output_csv_path_interesados}")

    print(f"Archivo Excel MTJ guardado exitosamente en: {output_file_path_mtj}")
    print(f"Archivo CSV MTJ guardado exitosamente en: {output_csv_path_mtj}")
    print(f"Archivo Excel INTERESADOS guardado exitosamente en: {output_file_path_interesados}")
    print(f"Archivo CSV INTERESADOS guardado exitosamente en: {output_csv_path_interesados}")
