import os
import pandas as pd
import csv
import datetime
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from PyQt5.QtWidgets import (
    QVBoxLayout, QLabel, QPushButton, QLineEdit, QCheckBox, QTextEdit, QFileDialog, QMessageBox, QDialog
)
from PyQt5.QtCore import QCoreApplication,Qt


def open_estandarizacion_window():
    # Crear un cuadro de diálogo para la estandarización
    estandarizacion_window = QDialog()
    estandarizacion_window.setWindowTitle("Estandarización masiva de MTJ")
    estandarizacion_window.setMinimumSize(600, 400)

    # Layout principal
    layout = QVBoxLayout()

    # Variables para las rutas
    input_folder_var = QLineEdit()
    output_folder_var = QLineEdit()

    # Variables para las opciones
    generar_csv_var = QCheckBox("Generar CSV individuales")
    procesar_interesados_var = QCheckBox("Agregar Procesamiento de la hoja INTERESADOS")
    unir_archivos_var = QCheckBox("Unir todos los archivos estandarizados en uno solo (estandariza excel individuales y genera consolidado)")
    solo_unificado_var = QCheckBox("Generar únicamente excel y csv consolidados de MTJ estandarizada")

    # Funciones para seleccionar carpetas
    def seleccionar_input_folder():
        input_folder = QFileDialog.getExistingDirectory(estandarizacion_window, "Seleccione la carpeta que contiene los archivos Excel a convertir")
        if input_folder:
            input_folder_var.setText(input_folder)

    def seleccionar_output_folder():
        output_folder = QFileDialog.getExistingDirectory(estandarizacion_window, "Seleccione la carpeta de salida")
        if output_folder:
            output_folder_var.setText(output_folder)

    # Configuración de la interfaz
    layout.addWidget(QLabel("Carpeta de entrada:"))
    layout.addWidget(input_folder_var)
    boton_seleccionar_input = QPushButton("Seleccionar")
    boton_seleccionar_input.clicked.connect(seleccionar_input_folder)
    layout.addWidget(boton_seleccionar_input)

    layout.addWidget(QLabel("Carpeta de salida:"))
    layout.addWidget(output_folder_var)
    boton_seleccionar_output = QPushButton("Seleccionar")
    boton_seleccionar_output.clicked.connect(seleccionar_output_folder)
    layout.addWidget(boton_seleccionar_output)

    layout.addWidget(generar_csv_var)
    layout.addWidget(procesar_interesados_var)
    layout.addWidget(unir_archivos_var)
    layout.addWidget(solo_unificado_var)

    # Cuadro de texto para mostrar el estado detallado del progreso
    status_text = QTextEdit()
    status_text.setReadOnly(True)
    layout.addWidget(status_text)

    def update_progress(message):
        status_text.append(message)
        status_text.verticalScrollBar().setValue(status_text.verticalScrollBar().maximum())
        QCoreApplication.processEvents()  # Asegurar que se procesen eventos pendientes

    def iniciar_procesamiento():
        input_folder = input_folder_var.text()
        output_folder = output_folder_var.text()

        # Verificar que las carpetas estén seleccionadas
        if not input_folder or not output_folder:
            QMessageBox.warning(estandarizacion_window, "Error", "Debe seleccionar las carpetas de entrada y salida.")
            return

        # Realizar la estandarización directamente (sin hilos)
        ejecutar_estandarizacion_masiva(
            input_folder, output_folder, generar_csv_var.isChecked(), procesar_interesados_var.isChecked(),
            unir_archivos_var.isChecked(), solo_unificado_var.isChecked(), update_progress
        )

    # Botón para iniciar el procesamiento
    boton_iniciar = QPushButton("Iniciar procesamiento")
    boton_iniciar.clicked.connect(iniciar_procesamiento)
    layout.addWidget(boton_iniciar)

    # Establecer el layout
    estandarizacion_window.setLayout(layout)
    estandarizacion_window.exec_()

def aplicar_formato_fecha(worksheet, df, columnas_fecha, date_style, update_progress):
    for col in columnas_fecha:
        if col in df.columns:
            col_idx = df.columns.get_loc(col) + 1  # Obtener el índice de la columna (1-indexed)
            column_name = df.columns[col_idx - 1]  # Obtener el nombre de la columna
            for row in range(2, len(df) + 2):  # Comenzar desde la fila 2 hasta el final de los datos
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value is not None:
                    try:
                        # Convertir la celda a fecha si es posible
                        date_value = pd.to_datetime(cell.value).date()
                        cell.value = date_value
                        cell.number_format = 'DD/MM/YYYY'
                        cell.style = date_style
                    except (ValueError, OverflowError):
                        # Si falla la conversión, mantener el valor como texto y mostrar un mensaje detallado
                        original_value = cell.value
                        cell.value = str(cell.value)
                        cell.number_format = '@'  # Formato de texto
                        update_progress(f"Advertencia: El valor '{original_value}' en la celda {cell.coordinate} de la columna '{column_name}' no es una fecha válida y fue tratado como texto.")

def save_excel_as_csv(excel_path, csv_path, update_progress):
    # Cargar el archivo Excel
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    # Abrir el archivo CSV para escritura
    with open(csv_path, mode='w', newline='', encoding='utf-8-sig') as f:
        csv_writer = csv.writer(f, delimiter=';')
        
        # Iterar sobre las filas de la hoja Excel
        for row in sheet.iter_rows(values_only=True):
            # Formatear las celdas de fecha en "DD/MM/YYYY"
            row = [cell.strftime('%d/%m/%Y') if isinstance(cell, (datetime.date, datetime.datetime)) else cell for cell in row]
            csv_writer.writerow(row)
    
    # Actualizar el progreso para notificar que el archivo CSV fue guardado
    update_progress(f"Archivo CSV guardado en: {csv_path}")

def ejecutar_estandarizacion_masiva(input_folder_path, output_folder, generar_csv, procesar_interesados, unir_archivos, solo_unificado, update_progress):
    try:
        update_progress(f"Procesando archivos en: {input_folder_path}")

        input_files = [f for f in os.listdir(input_folder_path) if f.endswith('.xlsx')]
        all_dfs_mtj = []
        all_dfs_interesados = []

        # Definir la lista de columnas de fecha al inicio
        date_columns_mtj = ['fecha_asig', 'fecha_analisi_tecnico', 'fecha_analisi_juridico', 'fmi_fecha_documento_fuente',
                            'fecha_analisi_agro', 'fecha_doc_acto_apertura', 'fecha_consulta_catastral']
        
        date_columns_interesados = ['fecha_inicio_tenencia']

        # Definir el estilo de fecha globalmente
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        
        for file_name in input_files:
            input_file_path = os.path.join(input_folder_path, file_name)
            update_progress(f"Procesando archivo: {file_name}")

            df_mtj = pd.read_excel(input_file_path, sheet_name='MTJ', header=4)
            if procesar_interesados:
                df_interesados = pd.read_excel(input_file_path, sheet_name='1_INTERESADOS', header=0).iloc[:, :18]

            max_columns_mtj = 124
            if df_mtj.shape[1] > max_columns_mtj:
                df_mtj = df_mtj.iloc[:, :max_columns_mtj]

            new_headers_mtj = [
                "ID", "uit", "local_id", "id_operacion", "fecha_asig", "profesional_tecnico", "fecha_analisi_tecnico", "estado_tecnico", 
                "observaciones_tecnico", "CR_InteresadoTipo", "CR_DocumentoTipo", "CR_Interesado_all", "Documento_Identidad", "jefe_hogar", 
                "victima_conflicto", "estado_civil", "inicio_tenencia", "reside_predio", "persona_distinta_reside", "quien", "explota_predio", 
                "metodo_levan", "area_terreno_levantamiento", "nombre_predio", "destino_economico", "tiene_rezago", "num_rezago", 
                "area_lpp_resgitral", "area_dif_catas_resgi", "tiene_construcciones", "cons_cantidad", "cons_area", "departamento", 
                "municipio", "vereda", "numero_predial", "numero_predial_provisional", "propietario_catstral", "direccion_catastral", 
                "area_catastral", "area_catastral_poligono", "area_cons_catastral", "destino_economico_catastral", "fecha_consulta_catastral", 
                "tipo_novedad", "complemneto_tipo_novedad", "clasificacion_suelo", "uso_suelo_plan_parcial", "categ_suelo_rural", 
                "categoria_proteccion", "desarrollo_restringido", "fecha_certificacion_uso_suelo", "cruza_determinantes", "cruza_desc_restricciones", 
                "cruza_desc_restricciones_area_por", "cruza_desc_condicionantes", "cruza_desc_condicionantes_area_por", "cruza_area_habilitada", 
                "concepto_catastral_general", "concepto_catastral_especifico", "req_acta_colindancia", "req_uso_suelo", "res_certificado_riesgos", 
                "profesional_juridico", "fecha_analisi_juridico", "estado_juridico", "observaciones_juridico", "formal_informal", "naturaleza_predio", 
                "tipo_predio", "tipo_derecho", "condicion_predio", "relaciona_fmi", "Codigo_Orip", "Matricula_Inmobiliaria", 
                "ExtReferenciaRegistralSistemaAntiguo", "fmi_propietario", "fmi_ref_catastral", "fmi_area", "fmi_direccion", "fmi_tipo_predio", 
                "fmi_matriz", "fmi_derivados", "tipo_fuente_adm", "fmi_ente_emisor", "numero_fuente_adm", "fmi_fecha_documento_fuente", 
                "disponibilidad_fuente_adm", "documento_apertura_fmi", "fecha_doc_acto_apertura", "fmi_medidas_cautelares", "fmi_salvedades_correc", 
                "fmi_complementa", "fmi_cabida_lindero", "fmi_titulo_originario", "fmi_estado", "tipo_novedad_fmi", "estado_rtdaf", "rupta", 
                "predio_cruza_pretension", "tipo_pretension", "medida_2333", "fadc", "req_ospr", "concepto_juridico", "tipologia_titulacion", 
                "ruta_atencion", "fmi_procedimientos_catatrales", "material_cimple", "doc_faltantes", "habilitado_reso", "reso", "profesional_agro", 
                "fecha_analisi_agro", "estado_agro", "observaciones_agro", "agro_metodo_analisis", "agro_ufh", "agro_cober_agro", "agrp_fuente_info", 
                "agro_aptitud", "agro_cultivos_ilicitos", "agro_recomendaciones", "agro_concepto"
            ]

            new_headers_interesados = [
                "qr_derivado", "interesado_tipo", "documento_tipo", "primer_nombre", "segundo_nombre", "primer_apellido", "segundo_apellido", 
                "interesado_concat", "documento_numero", "jefe_hogar", "victima_conflicto", "estado_civil", "fecha_inicio_tenencia", 
                "reside_predio_interesado", "reside_distinta_interesado", "quien", "explota_directamente", "numero_formulario_reso"
            ]

            df_mtj.columns = new_headers_mtj
            if procesar_interesados:
                df_interesados.columns = new_headers_interesados

            columns_to_replace = ['CR_InteresadoTipo', 'CR_DocumentoTipo', 'CR_Interesado_all', 'Documento_Identidad']
            for col in columns_to_replace:
                if col in df_mtj.columns:
                    df_mtj[col] = df_mtj[col].str.replace(';', ',')

            df_mtj = df_mtj.dropna(subset=['local_id', 'id_operacion'])

            if procesar_interesados:
                df_interesados = df_interesados.dropna(subset=['qr_derivado', 'interesado_tipo'])

            df_mtj['ID'] = pd.to_numeric(df_mtj['ID'], errors='coerce')
            max_id = df_mtj['ID'].max()
            df_mtj['ID'] = df_mtj['ID'].fillna(max_id + 1).astype(int)

            if not solo_unificado:
                base_filename = os.path.splitext(file_name)[0]
                output_file_path_mtj = os.path.join(output_folder, f"{base_filename}_MTJ_ESTANDARIZADA.xlsx")
                df_mtj.to_excel(output_file_path_mtj, index=False)

                if procesar_interesados:
                    output_file_path_interesados = os.path.join(output_folder, f"{base_filename}_INTERESADOS_ESTANDARIZADO.xlsx")
                    df_interesados.to_excel(output_file_path_interesados, index=False)

                # Cargar los workbooks para aplicar los estilos de fecha
                workbook_mtj = load_workbook(output_file_path_mtj)
                worksheet_mtj = workbook_mtj.active

                if 'date_style' not in workbook_mtj.named_styles:
                    workbook_mtj.add_named_style(date_style)

                # Aplicar el estilo de fecha corta a las columnas específicas en MTJ
                aplicar_formato_fecha(worksheet_mtj, df_mtj, date_columns_mtj, date_style, update_progress)
                workbook_mtj.save(output_file_path_mtj)

                if procesar_interesados:
                    workbook_interesados = load_workbook(output_file_path_interesados)
                    worksheet_interesados = workbook_interesados.active
                    aplicar_formato_fecha(worksheet_interesados, df_interesados, date_columns_interesados, date_style, update_progress)
                    workbook_interesados.save(output_file_path_interesados)

                if generar_csv:
                    output_csv_path_mtj = os.path.join(output_folder, f"{base_filename}_MTJ_ESTANDARIZADA.csv")
                    save_excel_as_csv(output_file_path_mtj, output_csv_path_mtj, update_progress)
                    
                    if procesar_interesados:
                        output_csv_path_interesados = os.path.join(output_folder, f"{base_filename}_INTERESADOS_ESTANDARIZADO.csv")
                        save_excel_as_csv(output_file_path_interesados, output_csv_path_interesados, update_progress)

            all_dfs_mtj.append(df_mtj)
            if procesar_interesados:
                all_dfs_interesados.append(df_interesados)

        # Consolidar archivos si unir_archivos o solo_unificado están activos
        if unir_archivos or solo_unificado:
            consolidated_df_mtj = pd.concat(all_dfs_mtj, ignore_index=True)
            consolidated_output_path_mtj = os.path.join(output_folder, "CONSOLIDADO_MTJ_ESTANDARIZADO.xlsx")
            consolidated_df_mtj.to_excel(consolidated_output_path_mtj, index=False)
            update_progress(f"Archivo consolidado guardado en: {consolidated_output_path_mtj}")

            # Aplicar formato de fecha al archivo consolidado MTJ
            workbook_consolidated_mtj = load_workbook(consolidated_output_path_mtj)
            worksheet_consolidated_mtj = workbook_consolidated_mtj.active
            aplicar_formato_fecha(worksheet_consolidated_mtj, consolidated_df_mtj, date_columns_mtj, date_style, update_progress)
            workbook_consolidated_mtj.save(consolidated_output_path_mtj)

            # Generar CSV para el consolidado MTJ
            consolidated_csv_path_mtj = os.path.join(output_folder, "CONSOLIDADO_MTJ_ESTANDARIZADO.csv")
            save_excel_as_csv(consolidated_output_path_mtj, consolidated_csv_path_mtj, update_progress)

            if procesar_interesados:
                consolidated_df_interesados = pd.concat(all_dfs_interesados, ignore_index=True)
                consolidated_output_path_interesados = os.path.join(output_folder, "CONSOLIDADO_INTERESADOS_ESTANDARIZADO.xlsx")
                consolidated_df_interesados.to_excel(consolidated_output_path_interesados, index=False)
                update_progress(f"Archivo consolidado guardado en: {consolidated_output_path_interesados}")

                # Aplicar formato de fecha al archivo consolidado INTERESADOS
                workbook_consolidated_interesados = load_workbook(consolidated_output_path_interesados)
                worksheet_consolidated_interesados = workbook_consolidated_interesados.active
                aplicar_formato_fecha(worksheet_consolidated_interesados, consolidated_df_interesados, date_columns_interesados, date_style, update_progress)
                workbook_consolidated_interesados.save(consolidated_output_path_interesados)

        update_progress("Procesamiento completado.")
  

        # Agregar mensaje de finalización con el enlace a la carpeta de salida
        dialog = QDialog()
        dialog.setWindowTitle("Estandarización completada")

        # Layout del diálogo
        layout = QVBoxLayout()

        # Texto del mensaje con enlace para abrir la carpeta
        label = QLabel(f"La estandarización se completó con éxito.<br>"
                       f"<a href='file:///{output_folder}'>Abrir carpeta</a>")
        label.setTextFormat(Qt.RichText)
        label.setTextInteractionFlags(Qt.TextBrowserInteraction)
        label.setOpenExternalLinks(True)

        # Botón Aceptar
        button_aceptar = QPushButton("Aceptar")
        button_aceptar.clicked.connect(dialog.accept)

        # Añadir los widgets al layout
        layout.addWidget(label)
        layout.addWidget(button_aceptar)

        dialog.setLayout(layout)
        dialog.exec_()

    except Exception as e:
        update_progress(f"Error durante la estandarización: {e}")